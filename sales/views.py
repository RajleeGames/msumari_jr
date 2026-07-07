# sales/views.py
from decimal import Decimal
import json
from datetime import date, datetime, timedelta
from django.db.models import F, ExpressionWrapper, DecimalField
from django.shortcuts import render
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (
    Sum, Q, OuterRef, Subquery, IntegerField, Value,
    F, ExpressionWrapper, DecimalField
)
from django.db.models.functions import TruncMonth, TruncYear, TruncDate, TruncWeek, Coalesce
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.shortcuts import redirect, get_object_or_404
from django.template.loader import select_template
from django.utils.timezone import localdate
from django.utils.dateformat import format as datefmt

from users.utils import role_required
from users.models import Role
from .models import MyDebt, MyDebtPayment
from .forms import MyDebtForm, MyDebtPaymentForm
from inventory.models import Product, Category, StockEntry, Supplier, Branch, ProductStock
from .models import Invoice, InvoiceItem, Customer, Payment, Expense
from .models import GenjiSale
from .forms import GenjiSaleForm
from django.apps import apps
from collections import OrderedDict

from decimal import Decimal
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Value
from django.db.models.functions import Coalesce
from inventory.models import ProductStock

# ───────────────────────────────────────────────────────────────
# TEMPLATE RENDER HELPER (supports both "sales/x.html" and "x.html")
# ───────────────────────────────────────────────────────────────
def _render(request, template_candidates, context):
    """
    Example:
      _render(request, ["sales/invoice_detail.html", "invoice_detail.html"], {...})
    """
    tpl = select_template(template_candidates)
    return HttpResponse(tpl.render(context, request))


# ───────────────────────────────────────────────────────────────
# ROLE / PROFILE / BRANCH HELPERS
# ───────────────────────────────────────────────────────────────
def _get_user_profile(request):
    return getattr(request.user, "profile", None)


def _role_lower(request):
    profile = _get_user_profile(request)
    return (getattr(profile, "role", "") or "").lower()


def _role_const(name, fallback):
    return str(getattr(Role, name, fallback)).lower()


def _is_admin(request):
    profile = _get_user_profile(request)
    return bool(request.user.is_superuser or (profile and str(getattr(profile, "role", "")).lower() == _role_const("ADMIN", "admin")))


def _is_cashier(request):
    return _role_lower(request) == _role_const("CASHIER", "cashier")


def _is_seller(request):
    return _role_lower(request) == _role_const("SELLER", "seller")


def _get_user_branch(request):
    """
    Expect: request.user.profile.branch (FK to inventory.Branch)
    If not present -> None
    """
    profile = _get_user_profile(request)
    if not profile:
        return None

    branch = getattr(profile, "branch", None)
    if branch:
        return branch

    branch_id = getattr(profile, "branch_id", None)
    if branch_id:
        return Branch.objects.filter(pk=branch_id).first()

    return None


# ───────────────────────────────────────────────────────────────
# SAFE UTILITIES
# ───────────────────────────────────────────────────────────────
def _safe_int(v, default=None):
    try:
        v = (v or "").strip() if isinstance(v, str) else v
        return int(v) if v not in (None, "") else default
    except Exception:
        return default


def _money(val) -> Decimal:
    try:
        return Decimal(val or 0)
    except Exception:
        return Decimal("0.00")


def _product_selling_price(prod) -> Decimal:
    # supports new field selling_price, fallback old price
    if hasattr(prod, "selling_price") and prod.selling_price is not None:
        return Decimal(prod.selling_price)
    return Decimal(getattr(prod, "price", 0) or 0)


def _product_buying_price(prod) -> Decimal:
    # supports new field buying_price, fallback 0
    if hasattr(prod, "buying_price") and prod.buying_price is not None:
        return Decimal(prod.buying_price)
    return Decimal("0.00")


# ───────────────────────────────────────────────────────────────
# BRANCH STOCK HELPERS
# ───────────────────────────────────────────────────────────────
def _branch_products_qs(branch, cat=None, q=None):
    """
    Returns products visible to seller for a branch with branch_qty annotation.
    If branch is None -> fallback to Product.stock.
    """
    if not branch:
        qs = Product.objects.select_related("category")
        if cat:
            qs = qs.filter(category_id=cat)
        if q:
            qs = qs.filter(name__icontains=q)
        qs = qs.filter(stock__gt=0)  # legacy mode
        return qs, None

    ps = ProductStock.objects.filter(
        product=OuterRef("pk"),
        branch=branch
    ).values("quantity")[:1]

    qs = (
        Product.objects
        .select_related("category")
        .annotate(branch_qty=Coalesce(Subquery(ps, output_field=IntegerField()), Value(0)))
        .filter(branch_qty__gt=0)
    )

    if cat:
        qs = qs.filter(category_id=cat)
    if q:
        qs = qs.filter(name__icontains=q)

    return qs, "branch_qty"


def _get_branch_qty(product_id, branch):
    if not branch:
        p = Product.objects.filter(pk=product_id).only("stock").first()
        return int(getattr(p, "stock", 0) or 0) if p else 0

    return int(
        ProductStock.objects.filter(product_id=product_id, branch=branch)
        .values_list("quantity", flat=True)
        .first() or 0
    )


def _adjust_branch_stock_locked(product, branch, delta):
    """
    Atomic safe change:
      delta negative for sale, positive for restock
    NOTE: In your system you prefer using StockEntry to update stock (recommended).
    """
    ps = ProductStock.objects.select_for_update().filter(product=product, branch=branch).first()
    if not ps:
        ps = ProductStock(product=product, branch=branch, quantity=0)

    new_qty = int(ps.quantity) + int(delta)
    if new_qty < 0:
        raise ValueError(f"Insufficient stock for '{product.name}' in {branch.name}. Available: {ps.quantity}")

    ps.quantity = new_qty
    ps.save()
    return new_qty


# ───────────────────────────────────────────────────────────────
# BRANCH DROPDOWN + SCOPE HELPERS
# ───────────────────────────────────────────────────────────────
def _branches_for_sales_dropdown():
    """
    Exclude Transport/HQ branches from sales dashboards.
    Adjust keywords to match your exact names.
    """
    return (
        Branch.objects.filter(is_active=True)
        .exclude(name__icontains="transport")
        .exclude(name__icontains="HAMA NA ")
        .exclude(name__icontains="hq")
        .order_by("name")
    )


def _can_filter_branch(request):
    # Admin + Cashier can use branch dropdown
    return _is_admin(request) or _is_cashier(request)


def _selected_branch(request):
    """
    Seller: always locked to their own branch
    Admin/Cashier: can pick ?branch=<id> (or None = all branches)
    """
    if _is_seller(request):
        return _get_user_branch(request)

    if not _can_filter_branch(request):
        return None

    branch_id = (request.GET.get("branch") or "").strip()
    if not branch_id:
        return None

    # Only allow branches from sales dropdown (not transport/hq)
    return _branches_for_sales_dropdown().filter(pk=branch_id).first()


def _invoice_scope_qs(request):
    """
    Base invoice scope:
    - seller => own branch only (if branch exists), else own invoices
    - admin/cashier => all
    """
    qs = Invoice.objects.select_related("customer", "seller").filter(invoice_state="active")
    if hasattr(Invoice, "branch_id"):
        qs = qs.select_related("branch")

    if _is_seller(request):
        b = _get_user_branch(request)
        if b and hasattr(Invoice, "branch_id"):
            qs = qs.filter(branch=b)
        else:
            qs = qs.filter(seller=request.user)

    return qs


def _expense_scope_qs(request):
    qs = Expense.objects.select_related("created_by")
    if hasattr(Expense, "branch_id"):
        qs = qs.select_related("branch")

    if _is_seller(request):
        b = _get_user_branch(request)
        if b and hasattr(Expense, "branch_id"):
            qs = qs.filter(branch=b)

    return qs


def _payment_scope_qs(request):
    qs = Payment.objects.select_related("invoice", "created_by")
    if hasattr(Invoice, "branch_id"):
        qs = qs.select_related("invoice__branch")

    if _is_seller(request):
        b = _get_user_branch(request)
        if b and hasattr(Invoice, "branch_id"):
            qs = qs.filter(invoice__branch=b)

    return qs


def _must_own_invoice_if_seller(request, inv):
    """
    Seller must only access own invoices / branch invoices.
    """
    if _is_admin(request) or _is_cashier(request):
        return True

    if _is_seller(request):
        branch = _get_user_branch(request)
        if branch and hasattr(inv, "branch_id"):
            return inv.branch_id == branch.id
        return inv.seller_id == request.user.id

    return False


from django.db.models import Sum
from .models import Payment, SalesReturn  # ✅ ensure SalesReturn imported


def _invoice_payments_total(inv) -> int:
    try:
        if hasattr(inv, "payments"):
            return int(inv.payments.aggregate(t=Sum("amount"))["t"] or 0)
    except Exception:
        pass
    return int(Payment.objects.filter(invoice=inv).aggregate(t=Sum("amount"))["t"] or 0)


def _invoice_returns_total(inv) -> int:
    try:
        if hasattr(inv, "returns"):
            qs = inv.returns.all()

            # count ONLY posted returns if model has status
            if hasattr(SalesReturn, "status"):
                qs = qs.filter(status__in=["posted", "POSTED", "Posted"])

            return int(qs.aggregate(t=Sum("total_amount"))["t"] or 0)
    except Exception:
        pass

    qs = SalesReturn.objects.filter(invoice=inv)
    if hasattr(SalesReturn, "status"):
        qs = qs.filter(status__in=["posted", "POSTED", "Posted"])
    return int(qs.aggregate(t=Sum("total_amount"))["t"] or 0)


def _invoice_balance(inv) -> int:
    """
    ✅ REAL debt balance after returns:
      balance = max(0, (total_amount - returns_total) - payments_total)

    Cancelled invoices = no debt balance
    """
    # cancelled invoice should not count as debt
    if (getattr(inv, "invoice_state", "active") or "active").lower() == "cancelled":
        return 0

    # only debt invoices use balance
    if (getattr(inv, "status", "") or "").lower() != "debt":
        return 0

    total_amount = int(getattr(inv, "total_amount", 0) or 0)
    returns_total = _invoice_returns_total(inv)
    payments_total = _invoice_payments_total(inv)

    effective_total = max(0, total_amount - returns_total)
    balance = max(0, effective_total - payments_total)

    return balance

def _parse_date_range(request):
    today = localdate()

    start_str = (request.GET.get("start_date") or "").strip()
    end_str = (request.GET.get("end_date") or "").strip()

    if not start_str and not end_str:
        start = today - timedelta(days=6)
        end = today
        return start, end, start.isoformat(), end.isoformat()

    try:
        start = datetime.strptime(start_str, "%Y-%m-%d").date()
    except Exception:
        start = today - timedelta(days=6)

    try:
        end = datetime.strptime(end_str, "%Y-%m-%d").date()
    except Exception:
        end = today

    if start > end:
        start, end = end, start

    return start, end, start.isoformat(), end.isoformat()


from collections import OrderedDict
from django.apps import apps


from collections import OrderedDict
from django.apps import apps


def _clean_contact_name(name):
    name = (name or "").strip()
    return name if name else "Customer"


def _clean_contact_phone(phone):
    return (phone or "").strip()


def _first_nonempty_attr(obj, candidates):
    """
    Return first non-empty attribute value from candidate field names.
    """
    for field in candidates:
        try:
            value = getattr(obj, field, None)
            if value is not None and str(value).strip():
                return str(value).strip()
        except Exception:
            pass
    return ""


def _get_model_safe(app_labels, model_names):
    """
    Try many app labels + model names safely.
    """
    for app_label in app_labels:
        for model_name in model_names:
            try:
                return apps.get_model(app_label, model_name)
            except Exception:
                pass
    return None


def _save_customer_contact(name=None, phone=None):
    """
    Save/update contact into Customer table.

    Rules:
    - if no name => default 'Customer'
    - if phone exists, use phone as main unique key
    - if no phone, match by name
    """
    raw_name = (name or "").strip()
    raw_phone = (phone or "").strip()

    # nothing entered at all
    if not raw_name and not raw_phone:
        return None

    clean_name = _clean_contact_name(raw_name)
    clean_phone = _clean_contact_phone(raw_phone)

    # phone is strongest match
    if clean_phone:
        obj = Customer.objects.filter(phone=clean_phone).first()
        if obj:
            old_name = (obj.name or "").strip().lower()
            weak_names = ["", "unknown", "customer", "walk-in", "n/a", "—"]
            if old_name in weak_names and clean_name:
                obj.name = clean_name
                obj.save(update_fields=["name"])
            return obj

        return Customer.objects.create(
            name=clean_name,
            phone=clean_phone
        )

    # fallback by name only
    obj = Customer.objects.filter(name__iexact=clean_name).first()
    if obj:
        return obj

    return Customer.objects.create(
        name=clean_name,
        phone=""
    )


def _collect_all_contacts():
    """
    ONLY collect contacts from:
    1. Customer table (sales/debt)
    2. GenjiSale
    3. Transport Trips
    4. Transport Bookings
    """
    contacts = OrderedDict()

    def add_contact(name="", phone="", source="System", obj_id=None):
        raw_name = (name or "").strip()
        raw_phone = (phone or "").strip()

        # skip if both are empty
        if not raw_name and not raw_phone:
            return

        clean_name = _clean_contact_name(raw_name)
        clean_phone = _clean_contact_phone(raw_phone)

        # dedupe by phone first, else by name
        key = f"phone:{clean_phone}" if clean_phone else f"name:{clean_name.lower()}"

        if key not in contacts:
            contacts[key] = {
                "id": obj_id or "",
                "name": clean_name,
                "phone": clean_phone or "",
                "source": source,
            }
        else:
            existing = contacts[key]

            weak_names = ["", "customer", "unknown", "walk-in", "n/a", "—"]
            if existing["name"].strip().lower() in weak_names and clean_name:
                existing["name"] = clean_name

            if not existing["phone"] and clean_phone:
                existing["phone"] = clean_phone

            if source and source not in existing["source"]:
                existing["source"] = f'{existing["source"]}, {source}'

    # --------------------------------------------------
    # 1) SALES / DEBT CUSTOMERS
    # --------------------------------------------------
    for c in Customer.objects.all().order_by("-id"):
        add_contact(
            name=getattr(c, "name", "") or "",
            phone=getattr(c, "phone", "") or "",
            source="Debt/Sales",
            obj_id=getattr(c, "id", "")
        )

    # --------------------------------------------------
    # 2) GENJI CONTACTS
    # --------------------------------------------------
    try:
        for g in GenjiSale.objects.all().order_by("-id"):
            g_name = _first_nonempty_attr(g, [
                "customer_name", "name", "client_name", "contact_name"
            ])
            g_phone = _first_nonempty_attr(g, [
                "customer_phone", "phone", "phone_number", "mobile", "contact_phone"
            ])

            add_contact(
                name=g_name,
                phone=g_phone,
                source="Genji",
                obj_id=getattr(g, "id", "")
            )
    except Exception:
        pass

    # --------------------------------------------------
    # 3) TRIP CONTACTS
    # --------------------------------------------------
    TripModel = _get_model_safe(
        app_labels=["transport", "trips"],
        model_names=["Trip", "Trips", "TransportTrip"]
    )

    if TripModel:
        try:
            for t in TripModel.objects.all().order_by("-id"):
                t_name = _first_nonempty_attr(t, [
                    "customer_name", "name", "passenger_name", "client_name",
                    "contact_name", "full_name", "receiver_name"
                ])
                t_phone = _first_nonempty_attr(t, [
                    "customer_phone", "phone", "phone_number", "mobile",
                    "contact_phone", "receiver_phone"
                ])

                add_contact(
                    name=t_name,
                    phone=t_phone,
                    source="Trip",
                    obj_id=getattr(t, "id", "")
                )
        except Exception:
            pass

    # --------------------------------------------------
    # 4) BOOKING CONTACTS
    # --------------------------------------------------
    BookingModel = _get_model_safe(
        app_labels=["transport", "bookings", "trips"],
        model_names=["Booking", "Bookings", "TransportBooking"]
    )

    if BookingModel:
        try:
            for b in BookingModel.objects.all().order_by("-id"):
                b_name = _first_nonempty_attr(b, [
                    "customer_name", "name", "passenger_name", "client_name",
                    "contact_name", "full_name", "receiver_name"
                ])
                b_phone = _first_nonempty_attr(b, [
                    "customer_phone", "phone", "phone_number", "mobile",
                    "contact_phone", "receiver_phone"
                ])

                add_contact(
                    name=b_name,
                    phone=b_phone,
                    source="Booking",
                    obj_id=getattr(b, "id", "")
                )
        except Exception:
            pass

    rows = list(contacts.values())
    rows.sort(key=lambda x: (0 if x["phone"] else 1, (x["name"] or "").lower()))
    return rows


# ───────────────────────────────────────────────────────────────
# LANDING REDIRECT "/"
# ───────────────────────────────────────────────────────────────
@login_required
def dashboard_redirect(request):
    role = _role_lower(request)

    if request.user.is_superuser or role == _role_const("ADMIN", "admin"):
        return redirect("sales:admin_dashboard")

    if role == _role_const("CASHIER", "cashier"):
        return redirect("sales:admin_dashboard")

    if role == _role_const("TRANSPORT", "transport"):
        return redirect("transport:index")

    if role == _role_const("SELLER", "seller"):
        return redirect("sales:seller_dashboard")

    return redirect("login")


# ───────────────────────────────────────────────────────────────
# SELLER POS (BRANCH-AWARE)
# ───────────────────────────────────────────────────────────────
@login_required
@role_required(["seller", "admin"])
def seller_dashboard(request):
    branch = _get_user_branch(request)

    raw_cart = request.session.get("cart", {}) or {}
    cart = {}

    # normalize cart (store as {"pid": {"qty": int, "discount": int}})
    for pid, entry in raw_cart.items():
        if isinstance(entry, dict):
            qty = max(1, int(entry.get("qty", 1) or 1))
            disc = max(0, int(entry.get("discount", 0) or 0))
        else:
            qty, disc = max(1, int(entry or 1)), 0
        cart[str(pid)] = {"qty": qty, "discount": disc}

    request.session["cart"] = cart
    request.session.modified = True

    # build cart items (✅ include unit_price so template never uses product.price)
    cart_items = []
    for pid, data in cart.items():
        prod = get_object_or_404(Product, pk=int(pid))
        qty = max(1, int(data.get("qty", 1) or 1))
        disc = max(0, int(data.get("discount", 0) or 0))  # per-unit discount

        unit_price = _product_selling_price(prod)  # Decimal
        discount_d = Decimal(str(disc))
        effective_price = unit_price - discount_d
        if effective_price < Decimal("0.00"):
            effective_price = Decimal("0.00")

        total_price = (effective_price * Decimal(qty)).quantize(Decimal("0.01"))

        # branch-aware visible stock
        if _is_seller(request) and branch:
            visible_stock = int(_get_branch_qty(prod.pk, branch) or 0)
        else:
            visible_stock = int(getattr(prod, "stock", 0) or 0)

        cart_items.append({
            "product": prod,
            "quantity": qty,
            "discount": disc,
            "unit_price": unit_price,            # ✅ USED IN TEMPLATE
            "effective_price": effective_price,  # optional (use in receipt/invoice if needed)
            "total_price": total_price,
            "visible_stock": visible_stock,
        })

    # filters (optional)
    cat = (request.GET.get("category") or "").strip() or None
    q = (request.GET.get("q") or "").strip() or None

    # branch products query
    qs, stock_field = _branch_products_qs(
        branch if _is_seller(request) else None,
        cat=cat,
        q=q
    )

    products_data = []
    for p in qs:
        stock_val = getattr(p, stock_field) if stock_field else getattr(p, "stock", 0)
        stock_val = int(stock_val or 0)

        price = _product_selling_price(p)  # Decimal

        # ✅ include "price" because your barcode JS expects pd.price
        products_data.append({
            "id": p.id,
            "name": p.name,
            "price": float(price),  # JS use
            "display": f"{p.name} — Tsh {int(price):,} (Stock: {stock_val})",
            "category": p.category_id,
            "stock": stock_val,
            "reorder_level": int(getattr(p, "reorder_level", 0) or 0),
        })

    return _render(
        request,
        ["sales/seller_dashboard.html", "seller_dashboard.html"],
        {
            "categories": Category.objects.values_list("id", "name"),
            "products_json": json.dumps(products_data),
            "cart_items": cart_items,
            "selected_cat": cat or "",
            "search_q": q or "",
            "branch": branch,
        },
    )


@login_required
@role_required(["seller", "admin"])
def add_to_cart(request, product_id):
    branch = _get_user_branch(request)
    qty = max(1, int(request.POST.get("quantity", 1)))

    if _is_seller(request) and branch:
        available = _get_branch_qty(product_id, branch)
        if available <= 0:
            messages.error(request, "No stock in your branch for this product.")
            return redirect("sales:seller_dashboard")
        if qty > available:
            qty = available
            messages.warning(request, f"Stock in your branch is only {available}. Quantity reduced.")

    cart = request.session.setdefault("cart", {})
    entry = cart.get(str(product_id), {"qty": 0, "discount": 0})
    entry["qty"] += qty
    cart[str(product_id)] = entry
    request.session.modified = True
    return redirect("sales:seller_dashboard")


@login_required
@role_required(["seller", "admin"])
def remove_from_cart(request, product_id):
    cart = request.session.get("cart", {})
    cart.pop(str(product_id), None)
    request.session.modified = True
    return redirect("sales:seller_dashboard")


@login_required
@role_required(["seller", "admin"])
def update_cart_qty(request, product_id):
    if request.method == "POST":
        branch = _get_user_branch(request)
        cart = request.session.setdefault("cart", {})

        if str(product_id) in cart:
            qty = max(1, int(request.POST.get("quantity", 1)))
            disc = max(0, int(request.POST.get("discount", 0)))

            if _is_seller(request) and branch:
                available = _get_branch_qty(product_id, branch)
                if available <= 0:
                    messages.error(request, "No stock in your branch for this product.")
                    cart.pop(str(product_id), None)
                    request.session.modified = True
                    return redirect("sales:seller_dashboard")
                if qty > available:
                    qty = available
                    messages.warning(request, f"Stock in your branch is only {available}. Quantity reduced.")

            cart[str(product_id)] = {"qty": qty, "discount": disc}
            request.session.modified = True

    return redirect("sales:seller_dashboard")


@login_required
@role_required(["seller", "admin"])
def add_selected_product_to_cart(request):
    if request.method == "POST":
        branch = _get_user_branch(request)

        pid = request.POST.get("product_id")
        if not pid:
            return redirect("sales:seller_dashboard")

        qty = max(1, int(request.POST.get("quantity", 1)))

        if _is_seller(request) and branch:
            available = _get_branch_qty(pid, branch)
            if available <= 0:
                messages.error(request, "No stock in your branch for this product.")
                return redirect("sales:seller_dashboard")
            if qty > available:
                qty = available
                messages.warning(request, f"Stock in your branch is only {available}. Quantity reduced.")

        cart = request.session.setdefault("cart", {})
        entry = cart.get(str(pid), {"qty": 0, "discount": 0})
        entry["qty"] += qty
        cart[str(pid)] = entry
        request.session.modified = True

    return redirect("sales:seller_dashboard")


@login_required
@role_required(["seller", "admin"])
def ajax_products(request):
    branch = _get_user_branch(request) if _is_seller(request) else None
    cat = (request.GET.get("category") or "").strip() or None
    q = (request.GET.get("q") or "").strip() or None

    qs, stock_field = _branch_products_qs(branch, cat=cat, q=q)

    results = []
    for p in qs[:50]:
        stock_val = getattr(p, stock_field) if stock_field else getattr(p, "stock", 0)
        stock_val = int(stock_val or 0)
        results.append({
            "id": p.id,
            "label": f"{p.name} — Tsh{_product_selling_price(p):.2f} (Stock: {stock_val})"
        })

    return JsonResponse(results, safe=False)


@login_required
@role_required(["seller", "admin"])
def lookup_product(request):
    branch = _get_user_branch(request) if _is_seller(request) else None
    barcode = (request.GET.get("barcode") or "").strip()

    try:
        prod = Product.objects.get(barcode=barcode)
    except Product.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Product not found"})

    stock_val = _get_branch_qty(prod.pk, branch) if branch else int(getattr(prod, "stock", 0) or 0)

    return JsonResponse({
        "status": "success",
        "product": {
            "id": prod.pk,
            "name": prod.name,
            "price": str(_product_selling_price(prod)),
            "stock": stock_val,
        }
    })

@login_required
@role_required(["seller", "admin"])
def submit_sale(request):
    """
    Seller handles everything:
      - cash/bank/ebt sale => paid=True immediately
      - debt => paid=False and appears in debts list
      - stock updates ONLY via StockEntry
      - branch-aware: seller sells only from their branch
      - customer name/phone allowed for ALL payment methods
      - saved contacts also go to Customer table
    """
    if request.method != "POST":
        return redirect("sales:seller_dashboard")

    cart = request.session.get("cart", {})
    if not cart:
        messages.warning(request, "Cart is empty.")
        return redirect("sales:seller_dashboard")

    branch = _get_user_branch(request) if _is_seller(request) else None

    payment = (request.POST.get("payment_method") or "debt").strip().lower()
    customer_name = (request.POST.get("customer_name") or "").strip()
    customer_phone = (request.POST.get("customer_phone") or "").strip()

    if payment not in ("cash", "debt", "ebt", "bank"):
        payment = "debt"

    # debt can require at least some customer identification
    if payment == "debt" and not customer_name and not customer_phone:
        messages.error(request, "For debt sale, enter at least customer name or phone number.")
        return redirect("sales:seller_dashboard")

    customer = None

    # Save customer for ALL payment methods if any info entered
    if customer_name or customer_phone:
        customer = _save_customer_contact(customer_name, customer_phone)

    # For debt, force customer creation even if one field missing
    if payment == "debt" and customer is None:
        customer = _save_customer_contact(customer_name or "Customer", customer_phone or "")

    # IMPORTANT: paid flag
    paid_flag = payment in ("cash", "bank", "ebt")

    try:
        with transaction.atomic():
            # 1) Validate stock first
            if branch:
                for pid, data in cart.items():
                    qty = max(1, int(data.get("qty", 1)))
                    prod = get_object_or_404(Product, pk=int(pid))

                    ps = (
                        ProductStock.objects
                        .select_for_update()
                        .filter(product=prod, branch=branch)
                        .first()
                    )
                    available = int(ps.quantity) if ps else 0

                    if qty > available:
                        raise ValueError(
                            f"Insufficient stock for '{prod.name}' in your branch. "
                            f"Available: {available}, Requested: {qty}"
                        )

            # 2) Create invoice
            inv_kwargs = {
                "customer": customer,
                "seller": request.user,
                "status": payment,
                "paid": paid_flag,
                "invoice_state": "active"
            }
            if branch and hasattr(Invoice, "branch_id"):
                inv_kwargs["branch"] = branch

            inv = Invoice.objects.create(**inv_kwargs)

            # 3) Create items + stock entries
            for pid, data in cart.items():
                prod = get_object_or_404(Product, pk=int(pid))
                qty = max(1, int(data.get("qty", 1)))
                disc = max(0, int(data.get("discount", 0)))

                selling = _product_selling_price(prod)
                buying = _product_buying_price(prod)

                item_kwargs = {
                    "invoice": inv,
                    "product": prod,
                    "quantity": qty,
                    "discount": disc,
                }

                if hasattr(InvoiceItem, "selling_price"):
                    item_kwargs["selling_price"] = selling
                    if hasattr(InvoiceItem, "buying_cost"):
                        item_kwargs["buying_cost"] = buying
                else:
                    item_kwargs["unit_price"] = selling

                InvoiceItem.objects.create(**item_kwargs)

                StockEntry.objects.create(
                    product=prod,
                    branch=branch if branch else None,
                    change=-qty,
                    note=f"Sale #{inv.pk} ({payment})",
                )

            # 4) Totals
            if hasattr(inv, "calculate_totals") and callable(getattr(inv, "calculate_totals")):
                inv.calculate_totals()

            # 5) Extra safety: if invoice has no customer but form had data, attach saved customer again
            if (customer_name or customer_phone) and getattr(inv, "customer", None) is None:
                customer = _save_customer_contact(customer_name, customer_phone)
                if customer:
                    inv.customer = customer
                    inv.save(update_fields=["customer"])

    except (ValueError, ValidationError) as e:
        messages.error(request, str(e))
        return redirect("sales:seller_dashboard")

    request.session["cart"] = {}
    request.session.modified = True
    messages.success(request, "Sale submitted successfully.")
    return redirect("sales:invoice_detail", pk=inv.pk)
# ───────────────────────────────────────────────────────────────

from decimal import Decimal
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from users.utils import role_required

from .models import Invoice, InvoiceItem  # adjust names if different





def _d(value, default="0.00"):
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal(default)


# sales/views.py
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from users.utils import role_required

from .models import Invoice, InvoiceItem, Payment


def _d(value, default="0.00"):
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal(default)


@login_required
@role_required(["seller", "admin", "cashier"])
def invoice_detail(request, pk=None, invoice_id=None):
    inv_id = pk if pk is not None else invoice_id
    invoice = get_object_or_404(Invoice, pk=inv_id)

    is_cancelled = (getattr(invoice, "invoice_state", "active") or "active").lower() == "cancelled"

    # role flags for template
    is_admin = False
    is_cashier = False
    is_seller = False

    try:
        from sales.views import _is_admin, _is_cashier, _is_seller
        is_admin = _is_admin(request)
        is_cashier = _is_cashier(request)
        is_seller = _is_seller(request)
    except Exception:
        role = ""
        try:
            role = (getattr(getattr(request.user, "profile", None), "role", "") or "").lower()
        except Exception:
            role = ""

        is_admin = role == "admin" or request.user.is_superuser
        is_cashier = role == "cashier"
        is_seller = role == "seller"

    can_cancel_invoice = (not is_cancelled) and (is_admin or is_cashier)

    # -----------------------------
    # ITEMS (safe computed list)
    # -----------------------------
    if hasattr(invoice, "items"):
        items_qs = invoice.items.select_related("product").all()
    else:
        items_qs = InvoiceItem.objects.filter(invoice=invoice).select_related("product")

    invoice_items = []
    total_items_profit = Decimal("0.00")
    total_items_discount = Decimal("0.00")

    for it in items_qs:
        qty = int(getattr(it, "quantity", 0) or 0)
        disc = _d(getattr(it, "discount", 0) or 0)

        prod = getattr(it, "product", None)

        # selling/unit price
        up = getattr(it, "unit_price", None)
        if up is None:
            up = getattr(it, "selling_price", None)
        if up is None and prod:
            up = getattr(prod, "selling_price", None)
        if up is None and prod:
            up = getattr(prod, "price", 0)
        unit_price = _d(up or 0)

        # buying price
        buy = getattr(it, "buying_cost", None)
        if buy is None and prod:
            buy = getattr(prod, "buying_price", None)
        if buy is None:
            buy = 0
        buying_price = _d(buy or 0)

        effective_price = unit_price - disc
        if effective_price < 0:
            effective_price = Decimal("0.00")

        line_total = effective_price * _d(qty)
        line_discount = disc * _d(qty)
        line_profit = (effective_price - buying_price) * _d(qty)

        total_items_profit += line_profit
        total_items_discount += line_discount

        invoice_items.append({
            "product": prod,
            "quantity": qty,
            "discount": disc,
            "unit_price": unit_price,
            "buying_price": buying_price,
            "effective_price": effective_price,
            "line_total": line_total,
            "line_discount": line_discount,
            "line_profit": line_profit,
        })

    # -----------------------------
    # PAYMENTS
    # -----------------------------
    if hasattr(invoice, "payments"):
        payments = invoice.payments.select_related("created_by").order_by("-timestamp")
        payments_total_int = int(invoice.payments.aggregate(t=Sum("amount"))["t"] or 0)
    else:
        payments = Payment.objects.filter(invoice=invoice).select_related("created_by").order_by("-timestamp")
        payments_total_int = int(Payment.objects.filter(invoice=invoice).aggregate(t=Sum("amount"))["t"] or 0)

    # -----------------------------
    # RETURNS
    # -----------------------------
    returns = []
    returns_total = Decimal("0.00")
    try:
        if hasattr(invoice, "returns"):
            returns = invoice.returns.select_related("created_by").order_by("-created_at")
            returns_total = _d(invoice.returns.aggregate(t=Sum("total_amount"))["t"] or 0)
    except Exception:
        returns = []
        returns_total = Decimal("0.00")

    # -----------------------------
    # EFFECTIVE TOTAL + BALANCE
    # -----------------------------
    total_amount = _d(getattr(invoice, "total_amount", 0) or 0)
    total_effective = total_amount - returns_total
    if total_effective < 0:
        total_effective = Decimal("0.00")

    effective_total_int = int(total_effective.to_integral_value())

    # cancelled invoice should always show 0 balance
    if is_cancelled:
        balance_value = 0
    else:
        balance_value = max(0, effective_total_int - payments_total_int)

        # only debt invoices use balance
        if (getattr(invoice, "status", "") or "").lower() != "debt":
            balance_value = 0

    return render(request, "sales/invoice_detail.html", {
        "invoice": invoice,
        "invoice_items": invoice_items,
        "payments": payments,
        "returns": returns,

        "returns_total": int(returns_total.to_integral_value()),
        "total_effective": effective_total_int,
        "payments_total": payments_total_int,
        "balance_value": balance_value,
        "is_cancelled": is_cancelled,

        # role flags
        "is_admin": is_admin,
        "is_cashier": is_cashier,
        "is_seller": is_seller,
        "can_cancel_invoice": can_cancel_invoice,

        # extra totals
        "total_items_profit": total_items_profit,
        "total_items_discount": total_items_discount,
    })
    
    
@login_required
@role_required(["seller", "admin", "cashier"])
def invoice_receipt(request, pk):
    inv = get_object_or_404(
        Invoice.objects.select_related("customer", "seller"),
        pk=pk
    )

    if not _must_own_invoice_if_seller(request, inv):
        return HttpResponseForbidden("You cannot access this invoice.")

    autoprint = (request.GET.get("autoprint") or "") == "1"
    is_cancelled = (getattr(inv, "invoice_state", "active") or "active").lower() == "cancelled"

    company = {
        "name": "MSUMARI JR STORE",
        "tin": "TIN: 129-816-910",
        "phone": "Tel: +255 XXX XXX XXX",
        "location": "Rau, Kazimoto Complex, Moshi",
        "footer": "THANK YOU FOR BUYING WITH US",
    }

    # payment method label
    status_map = {"cash": "CASH", "bank": "BANK", "ebt": "EBT", "debt": "DEBT"}
    status_label = status_map.get((inv.status or "").lower(), (inv.status or "").upper() or "—")

    # invoice state label
    invoice_state_label = "CANCELLED" if is_cancelled else "ACTIVE"

    # safe customer label
    customer_label = "Walk-in"
    if getattr(inv, "customer", None):
        customer_label = getattr(inv.customer, "name", None) or str(inv.customer)

    # safe payments total
    payments_total = 0
    try:
        payments_total = int(inv.payments.aggregate(t=Sum("amount"))["t"] or 0)
    except Exception:
        payments_total = 0

    # safe returns total
    returns_total = 0
    try:
        if hasattr(inv, "returns"):
            returns_total = int(inv.returns.aggregate(t=Sum("total_amount"))["t"] or 0)
    except Exception:
        returns_total = 0

    # safe total effective
    total_amount = int(getattr(inv, "total_amount", 0) or 0)
    total_effective = max(0, total_amount - returns_total)

    # cancelled invoice must show 0 balance
    def _safe_balance(i):
        if is_cancelled:
            return 0
        try:
            return int(_invoice_balance(i) or 0)
        except Exception:
            b = getattr(i, "balance", 0)
            try:
                v = b() if callable(b) else b
                return int(v or 0)
            except Exception:
                return 0

    return _render(request, ["sales/thermal_receipt.html", "thermal_receipt.html"], {
        "invoice": inv,
        "autoprint": autoprint,
        "company": company,
        "status_label": status_label,               # payment method
        "invoice_state_label": invoice_state_label, # ACTIVE / CANCELLED
        "is_cancelled": is_cancelled,
        "customer_label": customer_label,
        "payments_total": payments_total,
        "returns_total": returns_total,
        "total_effective": total_effective,
        "balance_value": _safe_balance(inv),
    })


@login_required
@role_required(["seller", "admin", "cashier"])
def outstanding_debts(request):
    q = (request.GET.get("q") or "").strip()
    selected = _selected_branch(request)

    # only ACTIVE debt invoices
    qs = _invoice_scope_qs(request).filter(status="debt", invoice_state="active")

    if selected and hasattr(Invoice, "branch_id") and _can_filter_branch(request):
        qs = qs.filter(branch=selected)

    if q:
        qs = qs.filter(customer__name__icontains=q)

    debts_rows = []
    for inv in qs.order_by("-created_at"):
        # _invoice_balance already returns 0 for cancelled invoices,
        # but we also filtered active only above for safety
        bal = _invoice_balance(inv)
        if bal > 0:
            debts_rows.append({
                "invoice": inv,
                "balance": bal,
                "is_cancelled": (getattr(inv, "invoice_state", "active") or "active").lower() == "cancelled",
            })

    total_outstanding = sum(r["balance"] for r in debts_rows)
    top_debtor = max(debts_rows, key=lambda x: x["balance"]) if debts_rows else None
    least_debtor = min(debts_rows, key=lambda x: x["balance"]) if debts_rows else None

    return _render(request, ["sales/outstanding_debts.html", "outstanding_debts.html"], {
        "debts": debts_rows,
        "query": q,
        "total_outstanding": total_outstanding,
        "top_debtor": top_debtor,
        "least_debtor": least_debtor,

        "branches": _branches_for_sales_dropdown() if _can_filter_branch(request) else [],
        "selected_branch": str(selected.id) if selected else "",
        "selected_branch_obj": selected,

        "is_admin": _is_admin(request),
        "is_cashier": _is_cashier(request),
        "is_seller": _is_seller(request),
        "branch": selected,
    })

@login_required
@role_required(["seller", "admin"])
def record_payment(request, invoice_id):
    inv = get_object_or_404(
    Invoice,
    pk=invoice_id,
    status="debt",
    invoice_state="active",
)

    if not _must_own_invoice_if_seller(request, inv):
        return HttpResponseForbidden("You cannot record payment for this invoice.")

    if request.method == "POST":
        raw = (request.POST.get("amount") or "0").strip()
        try:
            amount = int(Decimal(raw))
        except Exception:
            amount = 0

        if amount <= 0:
            messages.error(request, "Enter a valid amount.")
            return redirect("sales:record_payment", invoice_id=invoice_id)

        bal = _invoice_balance(inv)
        if amount > bal:
            amount = bal
            messages.warning(request, f"Amount reduced to balance: {bal} TSH")

        Payment.objects.create(
            invoice=inv,
            amount=amount,
            created_by=request.user
        )

        messages.success(request, "Payment recorded.")
        return redirect("sales:invoice_detail", pk=inv.pk)

    return _render(request, ["sales/record_payment.html", "record_payment.html"], {
        "invoice": inv,
        "balance": _invoice_balance(inv),
        "history": inv.payments.order_by("-timestamp"),
    })


def _product_cost_expr():
    """
    Picks the first existing buying-cost field from Product.
    Supported names (you can add yours here):
      buying_price, cost_price, purchase_price, buy_price, cost, unit_cost
    """
    candidates = ["buying_price", "cost_price", "purchase_price", "buy_price", "cost", "unit_cost"]
    product_field_names = {f.name for f in Product._meta.get_fields()}

    found = [c for c in candidates if c in product_field_names]
    if not found:
        # no cost field found
        return None, None

    # build Coalesce(product__field1, product__field2, ..., 0)
    out = DecimalField(max_digits=18, decimal_places=2)
    expr = F(f"product__{found[0]}")
    for name in found[1:]:
        expr = Coalesce(expr, F(f"product__{name}"))
    expr = Coalesce(expr, Value(0, output_field=out))
    return expr, found[0]


from decimal import Decimal
from django.db.models import Sum

def _dec(v):
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal("0.00")


def _pick_existing_field(model, candidates):
    fields = {f.name for f in model._meta.get_fields()}
    for c in candidates:
        if c in fields:
            return c
    return None


def compute_debt_stats(inv_qs, selected_branch=None, month=None, year=None):
    """
    ✅ Computes real debt after payments + returns

    BALANCE = (invoice.total_amount - returns_total) - payments_total

    Returns:
      num_debts               => count of invoices with balance > 0
      total_outstanding_debt  => sum of positive balances
      total_customer_credit   => sum of negative balances (as positive number)
    """
    from .models import Payment, SalesReturn

    debt_qs = inv_qs.filter(status="debt", invoice_state="active")

    # Optional filter (keep if you really want debt created in that month/year)
    if month and year:
        debt_qs = debt_qs.filter(created_at__month=month, created_at__year=year)

    invoice_ids = list(debt_qs.values_list("id", flat=True))
    if not invoice_ids:
        return 0, Decimal("0.00"), Decimal("0.00")

    # Invoice totals
    invoice_totals = {
        r["id"]: _dec(r["total_amount"])
        for r in debt_qs.values("id", "total_amount")
    }

    # Payments
    payment_map = {i: Decimal("0.00") for i in invoice_ids}
    pay_rows = Payment.objects.filter(invoice_id__in=invoice_ids).values("invoice_id").annotate(t=Sum("amount"))
    for r in pay_rows:
        payment_map[r["invoice_id"]] = _dec(r["t"] or 0)

    # Returns (pick correct field name safely)
    ret_amount_field = _pick_existing_field(SalesReturn, ["total_amount", "amount", "total", "value"])
    returns_map = {i: Decimal("0.00") for i in invoice_ids}

    if ret_amount_field:
        ret_qs = SalesReturn.objects.filter(invoice_id__in=invoice_ids)

        # branch filter
        if selected_branch:
            if hasattr(SalesReturn, "branch_id"):
                ret_qs = ret_qs.filter(branch=selected_branch)
            else:
                # if SalesReturn doesn't have branch, filter by invoice branch
                ret_qs = ret_qs.filter(invoice__branch=selected_branch)

        # posted filter (very flexible)
        if hasattr(SalesReturn, "status"):
            ret_qs = ret_qs.filter(status__iexact="posted") | ret_qs.filter(status__iexact="POSTED")

        ret_rows = ret_qs.values("invoice_id").annotate(t=Sum(ret_amount_field))
        for r in ret_rows:
            returns_map[r["invoice_id"]] = _dec(r["t"] or 0)

    # Calculate balances
    num_debts = 0
    total_outstanding = Decimal("0.00")
    total_credit = Decimal("0.00")

    for inv_id in invoice_ids:
        total = invoice_totals.get(inv_id, Decimal("0.00"))
        paid = payment_map.get(inv_id, Decimal("0.00"))
        returned = returns_map.get(inv_id, Decimal("0.00"))

        effective_total = total - returned
        balance = effective_total - paid  # ✅ can be negative

        if balance > 0:
            num_debts += 1
            total_outstanding += balance
        elif balance < 0:
            total_credit += (-balance)

    return num_debts, total_outstanding, total_credit

from decimal import Decimal
from datetime import date, datetime, timedelta
import json

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum
from django.shortcuts import render
from django.utils.timezone import localdate

from users.utils import role_required
from inventory.models import Product
from .models import Invoice, InvoiceItem, Payment, Expense


def _normalize_business_type(raw_value):
    value = (raw_value or "").strip().lower()
    allowed = {"electronics", "furniture", "magodoro"}
    return value if value in allowed else None


def _business_dashboard_title(business_type):
    return {
        "electronics": "Electronics Dashboard",
        "furniture": "Furniture Dashboard",
        "magodoro": "Magodoro Dashboard",
    }.get(business_type, "Admin Dashboard")

def compute_stock_worth(branch=None, business_type=None):
    """
    Stock worth = quantity * buying_price
    This shows only unsold stock value.
    """
    qs = ProductStock.objects.select_related("product", "branch")

    if branch:
        qs = qs.filter(branch=branch)

    if business_type:
        qs = qs.filter(product__business_type=business_type)

    worth_expr = ExpressionWrapper(
        F("quantity") * Coalesce(F("product__buying_price"), Value(0)),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )

    total = qs.aggregate(
        total=Coalesce(
            Sum(worth_expr),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )["total"] or Decimal("0.00")

    return Decimal(total)

def _business_dashboard_context(request, business_type=None):
    today = localdate()
    business_type = _normalize_business_type(business_type)

    # -------------------------------------------------
    # URL NAME FOR TEMPLATE QUICK FILTERS / RESET
    # -------------------------------------------------
    if business_type == "electronics":
        dashboard_url_name = "sales:electronics_dashboard"
    elif business_type == "furniture":
        dashboard_url_name = "sales:furniture_dashboard"
    elif business_type == "magodoro":
        dashboard_url_name = "sales:magodoro_dashboard"
    else:
        dashboard_url_name = "sales:admin_dashboard"

    # -------------------------------------------------
    # Branch filter
    # -------------------------------------------------
    branches_qs = _branches_for_sales_dropdown()
    branch_id = (request.GET.get("branch") or "").strip()
    selected_branch = branches_qs.filter(pk=branch_id).first() if branch_id else None

    inv_qs = Invoice.objects.filter(invoice_state="active")
    pay_qs = Payment.objects.select_related("invoice")
    exp_qs = Expense.objects.all()

    if selected_branch:
        if hasattr(Invoice, "branch_id"):
            inv_qs = inv_qs.filter(branch=selected_branch)
            pay_qs = pay_qs.filter(invoice__branch=selected_branch)
        if hasattr(Expense, "branch_id"):
            exp_qs = exp_qs.filter(branch=selected_branch)

    # -------------------------------------------------
    # DATE RANGE FILTER
    # default => today
    # -------------------------------------------------
    raw_from = (request.GET.get("date_from") or "").strip()
    raw_to = (request.GET.get("date_to") or "").strip()

    def _parse_ymd(v):
        try:
            return datetime.strptime(v, "%Y-%m-%d").date()
        except Exception:
            return None

    date_from = _parse_ymd(raw_from)
    date_to = _parse_ymd(raw_to)

    if not date_from and not date_to:
        date_from = today
        date_to = today
    elif date_from and not date_to:
        date_to = date_from
    elif date_to and not date_from:
        date_from = date_to

    if date_from > date_to:
        date_from, date_to = date_to, date_from

    is_today_range = (date_from == today and date_to == today)

    if is_today_range:
        range_label = "Today"
    elif date_from == date_to:
        range_label = date_from.strftime("%d %b %Y")
    else:
        range_label = f"{date_from.strftime('%d %b %Y')} → {date_to.strftime('%d %b %Y')}"

    # -------------------------------------------------
    # Period querysets
    # -------------------------------------------------
    period_inv_qs = inv_qs.filter(created_at__date__range=[date_from, date_to])
    period_pay_qs = pay_qs.filter(timestamp__date__range=[date_from, date_to])
    period_exp_qs = exp_qs.filter(date__range=[date_from, date_to])

    # -------------------------------------------------
    # BUSINESS TYPE FILTER
    # IMPORTANT:
    # invoices can contain mixed products,
    # so money must be grouped by InvoiceItem.product.business_type
    # -------------------------------------------------
    if business_type:
        period_items_qs = InvoiceItem.objects.filter(
            invoice__in=period_inv_qs,
            product__business_type=business_type,
            invoice__invoice_state="active",
        ).select_related("product", "invoice")

        today_items_qs = InvoiceItem.objects.filter(
            invoice__in=inv_qs.filter(created_at__date=today),
            product__business_type=business_type,
            invoice__invoice_state="active",
        ).select_related("product", "invoice")

        all_items_qs = InvoiceItem.objects.filter(
            invoice__in=inv_qs,
            product__business_type=business_type,
            invoice__invoice_state="active",
        ).select_related("product", "invoice")
    else:
        period_items_qs = InvoiceItem.objects.filter(
            invoice__in=period_inv_qs,
            invoice__invoice_state="active",
        ).select_related("product", "invoice")

        today_items_qs = InvoiceItem.objects.filter(
            invoice__in=inv_qs.filter(created_at__date=today),
            invoice__invoice_state="active",
        ).select_related("product", "invoice")

        all_items_qs = InvoiceItem.objects.filter(
            invoice__in=inv_qs,
            invoice__invoice_state="active",
        ).select_related("product", "invoice")

    def _sum_line_total(qs, payment_status=None, paid_only=False):
        if payment_status:
            qs = qs.filter(invoice__status=payment_status)
        if paid_only:
            qs = qs.filter(invoice__paid=True)

        total = Decimal("0.00")
        for item in qs:
            try:
                total += Decimal(str(item.line_total or 0))
            except Exception:
                pass
        return total

    # -------------------------------------------------
    # CASHFLOW (selected period)
    # -------------------------------------------------
    total_cash_period = _sum_line_total(period_items_qs, payment_status="cash", paid_only=True)
    total_bank_period = _sum_line_total(period_items_qs, payment_status="bank", paid_only=True)
    total_ebt_period = _sum_line_total(period_items_qs, payment_status="ebt", paid_only=True)

    if business_type:
        business_invoice_ids = list(
            period_items_qs.values_list("invoice_id", flat=True).distinct()
        )
        total_debt_payments = (
            period_pay_qs.filter(invoice_id__in=business_invoice_ids).aggregate(t=Sum("amount"))["t"]
            or Decimal("0.00")
        )
    else:
        total_debt_payments = (
            period_pay_qs.aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
        )

    expenses_period = (
        period_exp_qs.aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
    )

    total_received_period = (
        total_cash_period + total_bank_period + total_ebt_period + total_debt_payments
    )
    net_period_cashflow = total_received_period - expenses_period

    # -------------------------------------------------
    # TODAY SNAPSHOT
    # -------------------------------------------------
    total_cash_today = _sum_line_total(today_items_qs, payment_status="cash", paid_only=True)
    total_bank_today = _sum_line_total(today_items_qs, payment_status="bank", paid_only=True)
    total_ebt_today = _sum_line_total(today_items_qs, payment_status="ebt", paid_only=True)
    total_today = total_cash_today + total_bank_today + total_ebt_today

    expenses_today = (
        exp_qs.filter(date=today).aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
    )
    net_today_cashflow = total_today - expenses_today

    # -------------------------------------------------
    # OVERALL CASHFLOW
    # -------------------------------------------------
    overall_cash = _sum_line_total(all_items_qs, payment_status="cash", paid_only=True)
    overall_bank = _sum_line_total(all_items_qs, payment_status="bank", paid_only=True)
    overall_ebt = _sum_line_total(all_items_qs, payment_status="ebt", paid_only=True)

    if business_type:
        all_business_invoice_ids = list(
            all_items_qs.values_list("invoice_id", flat=True).distinct()
        )
        overall_debt_payments = (
            pay_qs.filter(invoice_id__in=all_business_invoice_ids).aggregate(t=Sum("amount"))["t"]
            or Decimal("0.00")
        )
    else:
        overall_debt_payments = pay_qs.aggregate(t=Sum("amount"))["t"] or Decimal("0.00")

    expenses_total = exp_qs.aggregate(t=Sum("amount"))["t"] or Decimal("0.00")

    overall_sales_cashflow = overall_cash + overall_bank + overall_ebt + overall_debt_payments
    net_overall_cashflow = overall_sales_cashflow - expenses_total

    # -------------------------------------------------
    # PROFIT
    # -------------------------------------------------
    cogs_period = Decimal("0.00")
    cogs_today = Decimal("0.00")
    revenue_period = Decimal("0.00")
    revenue_today = Decimal("0.00")

    for item in period_items_qs:
        try:
            revenue_period += Decimal(str(item.line_total or 0))
            qty = int(getattr(item, "quantity", 0) or 0)
            buying_cost = Decimal(str(getattr(item, "buying_cost", 0) or 0))
            cogs_period += (buying_cost * qty)
        except Exception:
            pass

    for item in today_items_qs:
        try:
            revenue_today += Decimal(str(item.line_total or 0))
            qty = int(getattr(item, "quantity", 0) or 0)
            buying_cost = Decimal(str(getattr(item, "buying_cost", 0) or 0))
            cogs_today += (buying_cost * qty)
        except Exception:
            pass

    profit_enabled = True
    cost_field_used = "buying_cost / product.buying_price snapshot"

    gross_profit_period = revenue_period - cogs_period
    net_profit_period = gross_profit_period - expenses_period

    gross_profit_today = revenue_today - cogs_today
    net_profit_today = gross_profit_today - expenses_today

    remaining_after_expenses_period = revenue_period - expenses_period
    remaining_after_expenses_today = revenue_today - expenses_today

    # -------------------------------------------------
    # STOCK WORTH
    # stock worth = quantity × buying_price
    # unsold stock only
    # -------------------------------------------------
    stock_qs = ProductStock.objects.select_related("product", "branch")

    if selected_branch:
        stock_qs = stock_qs.filter(branch=selected_branch)

    if business_type:
        stock_qs = stock_qs.filter(product__business_type=business_type)

    stock_worth_total = stock_qs.aggregate(
        total=Coalesce(
            Sum(
                ExpressionWrapper(
                    F("quantity") * Coalesce(F("product__buying_price"), Value(0)),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            ),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )["total"] or Decimal("0.00")

    # -------------------------------------------------
    # DEBTS
    # -------------------------------------------------
    if business_type:
        debt_invoice_ids = list(
            period_items_qs.filter(invoice__status="debt")
            .values_list("invoice_id", flat=True)
            .distinct()
        )
        debt_inv_qs = period_inv_qs.filter(id__in=debt_invoice_ids)
    else:
        debt_inv_qs = period_inv_qs

    num_debts, total_outstanding_debt, total_customer_credit = compute_debt_stats(
        debt_inv_qs,
        selected_branch=selected_branch,
    )

    # -------------------------------------------------
    # COUNTS (selected period)
    # -------------------------------------------------
    num_cash_sales = period_items_qs.filter(
        invoice__status="cash",
        invoice__paid=True
    ).values("invoice_id").distinct().count()

    num_bank_sales = period_items_qs.filter(
        invoice__status="bank",
        invoice__paid=True
    ).values("invoice_id").distinct().count()

    num_ebt_sales = period_items_qs.filter(
        invoice__status="ebt",
        invoice__paid=True
    ).values("invoice_id").distinct().count()

    total_transactions = period_items_qs.values("invoice_id").distinct().count()

    # -------------------------------------------------
    # RECENT PRODUCTS
    # -------------------------------------------------
    all_products = Product.objects.order_by("-id")
    if business_type:
        all_products = all_products.filter(business_type=business_type)

    paginator = Paginator(all_products, 6)
    recent_products = paginator.get_page(request.GET.get("recent_page", 1))

    # -------------------------------------------------
    # CHART 1: Daily sales trend
    # -------------------------------------------------
    sales_daily_map = {}

    for item in period_items_qs:
        day = item.invoice.created_at.date()
        sales_daily_map.setdefault(day, 0.0)
        try:
            sales_daily_map[day] += float(item.line_total or 0)
        except Exception:
            pass

    expenses_daily_raw = (
        period_exp_qs.values("date")
        .annotate(total=Sum("amount"))
        .order_by("date")
    )
    exp_daily_map = {r["date"]: float(r["total"] or 0) for r in expenses_daily_raw}

    chart_labels = []
    chart_sales = []
    chart_expenses = []

    cursor = date_from
    while cursor <= date_to:
        chart_labels.append(cursor.strftime("%d %b"))
        chart_sales.append(sales_daily_map.get(cursor, 0))
        chart_expenses.append(exp_daily_map.get(cursor, 0))
        cursor += timedelta(days=1)

    # -------------------------------------------------
    # CHART 2: Payment mix
    # -------------------------------------------------
    payment_mix_labels = ["Cash", "Bank", "EBT", "Debt Payments"]
    payment_mix_values = [
        float(total_cash_period),
        float(total_bank_period),
        float(total_ebt_period),
        float(total_debt_payments),
    ]

    # -------------------------------------------------
    # CHART 3: Last 6 months revenue vs expenses
    # -------------------------------------------------
    month_labels = []
    month_received_values = []
    month_expense_values = []

    month_cursor = date(today.year, today.month, 1)
    month_starts = []

    for offset in range(5, -1, -1):
        y = month_cursor.year
        m = month_cursor.month - offset

        while m <= 0:
            m += 12
            y -= 1
        while m > 12:
            m -= 12
            y += 1

        month_starts.append(date(y, m, 1))

    for mstart in month_starts:
        if mstart.month == 12:
            mend = date(mstart.year + 1, 1, 1) - timedelta(days=1)
        else:
            mend = date(mstart.year, mstart.month + 1, 1) - timedelta(days=1)

        month_labels.append(mstart.strftime("%b %Y"))

        month_inv_qs = inv_qs.filter(created_at__date__range=[mstart, mend])
        month_items_qs = InvoiceItem.objects.filter(
            invoice__in=month_inv_qs,
            invoice__invoice_state="active"
        )
        if business_type:
            month_items_qs = month_items_qs.filter(product__business_type=business_type)

        m_cash = _sum_line_total(month_items_qs, payment_status="cash", paid_only=True)
        m_bank = _sum_line_total(month_items_qs, payment_status="bank", paid_only=True)
        m_ebt = _sum_line_total(month_items_qs, payment_status="ebt", paid_only=True)

        if business_type:
            month_invoice_ids = list(month_items_qs.values_list("invoice_id", flat=True).distinct())
            m_debt_pay = (
                pay_qs.filter(timestamp__date__range=[mstart, mend], invoice_id__in=month_invoice_ids)
                .aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
            )
        else:
            m_debt_pay = (
                pay_qs.filter(timestamp__date__range=[mstart, mend])
                .aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
            )

        m_exp = (
            exp_qs.filter(date__range=[mstart, mend])
            .aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
        )

        month_received_values.append(float(m_cash + m_bank + m_ebt + m_debt_pay))
        month_expense_values.append(float(m_exp))

    context = {
        "is_admin": _is_admin(request),
        "is_cashier": _is_cashier(request),

        "branches": branches_qs,
        "selected_branch": str(selected_branch.id) if selected_branch else "",
        "selected_branch_obj": selected_branch,

        "business_type": business_type or "",
        "dashboard_title": _business_dashboard_title(business_type),
        "dashboard_url_name": dashboard_url_name,

        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "today_str": today.isoformat(),
        "last7_from": (today - timedelta(days=6)).isoformat(),
        "month_start_str": today.replace(day=1).isoformat(),
        "range_label": range_label,
        "is_today_range": is_today_range,

        "total_cash_period": total_cash_period,
        "total_bank_period": total_bank_period,
        "total_ebt_period": total_ebt_period,
        "total_received_period": total_received_period,
        "total_debt_payments": total_debt_payments,
        "expenses_period": expenses_period,
        "net_period_cashflow": net_period_cashflow,

        "total_cash_today": total_cash_today,
        "total_bank_today": total_bank_today,
        "total_ebt_today": total_ebt_today,
        "total_today": total_today,
        "expenses_today": expenses_today,
        "net_today_cashflow": net_today_cashflow,
        "net_overall_cashflow": net_overall_cashflow,

        "profit_enabled": profit_enabled,
        "profit_cost_field": cost_field_used or "",
        "revenue_today": revenue_today,
        "cogs_today": cogs_today,
        "gross_profit_today": gross_profit_today,
        "net_profit_today": net_profit_today,
        "revenue_period": revenue_period,
        "cogs_period": cogs_period,
        "gross_profit_period": gross_profit_period,
        "net_profit_period": net_profit_period,

        "remaining_after_expenses_period": remaining_after_expenses_period,
        "remaining_after_expenses_today": remaining_after_expenses_today,

        "stock_worth_total": stock_worth_total,

        "num_debts": num_debts,
        "total_outstanding_debt": total_outstanding_debt,
        "total_customer_credit": total_customer_credit,

        "num_cash_sales": num_cash_sales,
        "num_bank_sales": num_bank_sales,
        "num_ebt_sales": num_ebt_sales,
        "total_transactions": total_transactions,

        "recent_products": recent_products,

        "chart_labels_json": json.dumps(chart_labels),
        "chart_sales_json": json.dumps(chart_sales),
        "chart_expenses_json": json.dumps(chart_expenses),

        "payment_mix_labels_json": json.dumps(payment_mix_labels),
        "payment_mix_values_json": json.dumps(payment_mix_values),

        "month_labels_json": json.dumps(month_labels),
        "month_received_values_json": json.dumps(month_received_values),
        "month_expense_values_json": json.dumps(month_expense_values),
    }
    return context

@login_required
@role_required(["admin", "cashier"])
def admin_dashboard(request):
    context = _business_dashboard_context(request, business_type=None)
    return render(request, "sales/admin_dashboard.html", context)


@login_required
@role_required(["admin", "cashier"])
def electronics_dashboard(request):
    context = _business_dashboard_context(request, business_type="electronics")
    return render(request, "sales/admin_dashboard.html", context)


@login_required
@role_required(["admin", "cashier"])
def furniture_dashboard(request):
    context = _business_dashboard_context(request, business_type="furniture")
    return render(request, "sales/admin_dashboard.html", context)


@login_required
@role_required(["admin", "cashier"])
def magodoro_dashboard(request):
    context = _business_dashboard_context(request, business_type="magodoro")
    return render(request, "sales/admin_dashboard.html", context)

# ───────────────────────────────────────────────────────────────
# REPORTS (admin/cashier filter branch, seller locked to branch)
# ───────────────────────────────────────────────────────────────
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator

@login_required
@role_required(["admin", "cashier", "seller"])
def sales_history(request):
    selected = _selected_branch(request)
    sale_date = (request.GET.get("sale_date") or "").strip()

    qs = Invoice.objects.select_related("customer", "seller").prefetch_related("items__product")

    if selected and hasattr(Invoice, "branch_id") and _can_filter_branch(request):
        qs = qs.filter(branch=selected)

    if sale_date:
        qs = qs.filter(created_at__date=sale_date)

    qs = qs.order_by("-created_at")

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    invoice_rows = []

    for inv in page_obj.object_list:
        items_qs = (
            inv.items.select_related("product").all()
            if hasattr(inv, "items")
            else InvoiceItem.objects.filter(invoice=inv).select_related("product")
        )

        first_item = items_qs.first()
        first_product_name = "No items"
        more_items_count = 0
        items_count = 0

        if first_item:
            first_product_name = (
                getattr(getattr(first_item, "product", None), "name", "Unknown Product")
                or "Unknown Product"
            )
            items_count = items_qs.count()
            more_items_count = max(items_count - 1, 0)

        invoice_rows.append({
            "invoice": inv,
            "first_product_name": first_product_name,
            "items_count": items_count,
            "more_items_count": more_items_count,
        })

    return _render(request, ["sales/sales_history.html", "sales_history.html"], {
        "invoices": page_obj.object_list,
        "invoice_rows": invoice_rows,
        "page_obj": page_obj,
        "is_paginated": page_obj.has_other_pages(),
        "paginator": paginator,

        "branches": _branches_for_sales_dropdown() if _can_filter_branch(request) else [],
        "selected_branch": str(selected.id) if selected else "",
        "selected_branch_obj": selected,
        "sale_date": sale_date,

        "is_admin": _is_admin(request),
        "is_cashier": _is_cashier(request),
        "is_seller": _is_seller(request),
    })
    
from decimal import Decimal
from datetime import datetime, timedelta

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum
from django.utils.timezone import localdate

from .models import Invoice, Expense, Payment  # make sure these imports exist


def _dec(x):
    """
    Safe Decimal conversion. Keeps cents (does NOT cast to int).
    """
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal("0.00")


# ───────────────────────────────────────────────────────────────
# EXPENSES (ADD + HISTORY)  ✅ REQUIRED for urls.py
# ───────────────────────────────────────────────────────────────
@login_required
@role_required(["admin", "cashier", "seller"])
def add_expense(request):
    """
    URL: /expenses/add/
    Fixes: AttributeError: sales.views has no attribute add_expense
    """
    selected = _selected_branch(request)

    # Branch list for dropdown: only admin/cashier can pick, seller hidden
    branches = _branches_for_sales_dropdown() if _can_filter_branch(request) else []

    if request.method == "POST":
        # Inputs
        date_val = (request.POST.get("date") or "").strip()
        category = (request.POST.get("category") or "").strip()
        description = (request.POST.get("description") or "").strip()
        amount_raw = (request.POST.get("amount") or "0").strip()

        try:
            amount = Decimal(amount_raw)
        except Exception:
            amount = Decimal("0")

        if amount <= 0:
            messages.error(request, "Enter a valid amount.")
            return redirect("sales:kilasi_add_expense")

        # Determine branch
        branch_obj = None

        # seller locked to own branch
        if _is_seller(request):
            branch_obj = _get_user_branch(request)

        # admin/cashier can choose in form, else uses selected filter branch
        else:
            branch_id = (request.POST.get("branch") or "").strip()
            if branch_id:
                branch_obj = _branches_for_sales_dropdown().filter(pk=branch_id).first()
            elif selected:
                branch_obj = selected

        # Create expense
        exp_kwargs = {
            "date": date_val or localdate(),
            "category": category or "General",
            "description": description,
            "amount": amount,
            "created_by": request.user,
        }
        if hasattr(Expense, "branch_id"):
            exp_kwargs["branch"] = branch_obj

        Expense.objects.create(**exp_kwargs)

        messages.success(request, "Expense added successfully.")
        return redirect("sales:expense_history")

    return _render(request, ["sales/kilasi_add_expense.html", "kilasi_add_expense.html"], {
        "branches": branches,
        "selected_branch": str(selected.id) if selected else "",
        "selected_branch_obj": selected,
        "today_str": localdate().isoformat(),
        "is_admin": _is_admin(request),
        "is_cashier": _is_cashier(request),
        "is_seller": _is_seller(request),
    })


@login_required
@role_required(["admin", "cashier", "seller"])
def expense_history(request):
    """
    URL: /expenses/history/
    """
    selected = _selected_branch(request)
    branches = _branches_for_sales_dropdown() if _can_filter_branch(request) else []

    qs = _expense_scope_qs(request).order_by("-date", "-id")

    # admin/cashier branch filter
    if selected and _can_filter_branch(request) and hasattr(Expense, "branch_id"):
        qs = qs.filter(branch=selected)

    page_obj = Paginator(qs, 30).get_page(request.GET.get("page", 1))

    total_amount = qs.aggregate(t=Sum("amount"))["t"] or Decimal("0.00")

    return _render(request, ["sales/kilasi_expense_history.html", "kilasi_expense_history.html"], {
        "expenses": page_obj.object_list,
        "page_obj": page_obj,
        "is_paginated": page_obj.has_other_pages(),
        "paginator": page_obj.paginator,

        "total_amount": total_amount,

        "branches": branches,
        "selected_branch": str(selected.id) if selected else "",
        "selected_branch_obj": selected,

        "is_admin": _is_admin(request),
        "is_cashier": _is_cashier(request),
        "is_seller": _is_seller(request),
    })


@login_required
@role_required(["admin", "cashier", "seller"])
def sales_report(request):
    today = localdate()
    selected_date = (request.GET.get("date") or today.isoformat()).strip()

    try:
        report_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
    except ValueError:
        report_date = today
        selected_date = today.isoformat()

    prev_date = (report_date - timedelta(days=1)).isoformat()
    next_date = (report_date + timedelta(days=1)).isoformat()

    # branch filter: seller locked, admin/cashier optional
    selected = _selected_branch(request)

    # base scopes (respect role permissions)
    inv_qs = _invoice_scope_qs(request).filter(invoice_state="active")
    exp_qs = _expense_scope_qs(request)
    pay_qs = _payment_scope_qs(request)  # ✅ IMPORTANT: scope payments too (seller => branch)

    # Optional status filter (if dashboard passes ?status=cash/bank/ebt/debt)
    status_filter = (request.GET.get("status") or "").strip().lower()
    if status_filter:
        inv_qs = inv_qs.filter(status__iexact=status_filter)

    # Apply branch filter for admin/cashier only
    # (seller already locked by _invoice_scope_qs/_payment_scope_qs/_expense_scope_qs)
    if selected and _can_filter_branch(request):
        if hasattr(Invoice, "branch_id"):
            inv_qs = inv_qs.filter(branch=selected)

        if hasattr(Expense, "branch_id"):
            exp_qs = exp_qs.filter(branch=selected)

        # ✅ payments must follow the same selected branch
        if hasattr(Invoice, "branch_id"):
            pay_qs = pay_qs.filter(invoice__branch=selected)

    # ----------------------------
    # DAILY DATA
    # ----------------------------
    sales_qs = inv_qs.filter(created_at__date=report_date).order_by("-created_at")
    expenses_qs = exp_qs.filter(date=report_date).order_by("-date", "-id")

    # ✅ Income received that day: paid cash/bank/ebt invoices
    income_sales = _dec(
        sales_qs.filter(status__in=["cash", "bank", "ebt"], paid=True)
        .aggregate(t=Sum("total_amount"))["t"] or 0
    )

    # ✅ Debt payments received that day (scoped properly)
    income_debt = _dec(
        pay_qs.filter(timestamp__date=report_date)
        .aggregate(t=Sum("amount"))["t"] or 0
    )

    # ✅ Total revenue received
    revenue_received = income_sales + income_debt

    # ✅ expenses and net
    total_expenses = _dec(expenses_qs.aggregate(t=Sum("amount"))["t"] or 0)
    net_income = revenue_received - total_expenses

    # Pagination
    sales_page = Paginator(sales_qs, 20).get_page(request.GET.get("sales_page", 1))
    exp_page = Paginator(expenses_qs, 20).get_page(request.GET.get("exp_page", 1))

    return _render(
        request,
        ["sales/kilasi_sales_report.html", "kilasi_sales_report.html"],
        {
            "selected_date": selected_date,
            "prev_date": prev_date,
            "next_date": next_date,

            # template top cards
            "total_sales": revenue_received,
            "total_expenses": total_expenses,
            "net_sales": net_income,

            # ✅ breakdown block
            "revenue_received": revenue_received,
            "income_sales": income_sales,
            "income_debt": income_debt,

            "sales_page": sales_page,
            "exp_page": exp_page,

            "branches": _branches_for_sales_dropdown() if _can_filter_branch(request) else [],
            "selected_branch": str(selected.id) if selected else "",
            "selected_branch_obj": selected,
            "today_str": today.isoformat(),
        }
    )

@login_required
@role_required(["admin", "cashier", "seller"])
def enhanced_sales_report(request):
    today = localdate()
    selected = _selected_branch(request)

    date_filter = (request.GET.get("date") or "").strip()
    month_filter = _safe_int(request.GET.get("month"))
    year_filter = _safe_int(request.GET.get("year"))

    doughnut_month = _safe_int(request.GET.get("doughnut_month")) or today.month
    doughnut_year = _safe_int(request.GET.get("doughnut_year")) or today.year

    # ACTIVE invoices only
    inv_qs = _invoice_scope_qs(request).filter(invoice_state="active")
    exp_qs = _expense_scope_qs(request)
    pay_qs = _payment_scope_qs(request).filter(invoice__invoice_state="active")

    if selected and hasattr(Invoice, "branch_id") and _can_filter_branch(request):
        inv_qs = inv_qs.filter(branch=selected)
        pay_qs = pay_qs.filter(invoice__branch=selected)

    if selected and hasattr(Expense, "branch_id") and _can_filter_branch(request):
        exp_qs = exp_qs.filter(branch=selected)

    report_date = None
    if date_filter:
        try:
            report_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
        except ValueError:
            report_date = None

    if report_date:
        start = end = report_date
    else:
        end = today
        start = today - timedelta(days=6)

    # Daily chart
    daily_raw = (
        inv_qs.filter(created_at__date__gte=start, created_at__date__lte=end)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(total=Sum("total_amount"))
        .order_by("day")
    )
    daily_map = {r["day"]: float(r["total"] or 0) for r in daily_raw}
    days_count = (end - start).days + 1
    daily = [
        {
            "day": (start + timedelta(days=i)).isoformat(),
            "total": daily_map.get(start + timedelta(days=i), 0),
        }
        for i in range(days_count)
    ]

    # Weekly chart
    weekly_raw = (
        inv_qs.annotate(week=TruncWeek("created_at"))
        .values("week")
        .annotate(total=Sum("total_amount"))
        .order_by("week")
    )
    weeks = [r["week"] for r in weekly_raw if r["week"]]
    last12w = weeks[-12:] if len(weeks) >= 12 else weeks
    weekly_map = {r["week"]: float(r["total"] or 0) for r in weekly_raw}
    weekly = [{"week": w.strftime("%Y-%m-%d"), "total": weekly_map.get(w, 0)} for w in last12w]

    # Monthly chart (last 12 months)
    monthly = []
    cursor = date(today.year, today.month, 1)
    months_list = []
    for _ in range(12):
        months_list.append(cursor)
        prev_month = cursor.month - 1
        prev_year = cursor.year
        if prev_month == 0:
            prev_month = 12
            prev_year -= 1
        cursor = date(prev_year, prev_month, 1)
    months_list = list(reversed(months_list))

    monthly_raw = (
        inv_qs.annotate(m=TruncMonth("created_at"))
        .values("m")
        .annotate(total=Sum("total_amount"))
        .order_by("m")
    )
    monthly_map = {r["m"]: float(r["total"] or 0) for r in monthly_raw if r["m"]}

    for m in months_list:
        if month_filter and m.month != month_filter:
            continue
        if year_filter and m.year != year_filter:
            continue
        monthly.append({
            "month": m.month,
            "year": m.year,
            "total": monthly_map.get(m, 0),
        })

    # Yearly chart (last 5 years)
    years_list = [today.year - 4, today.year - 3, today.year - 2, today.year - 1, today.year]
    yearly_raw = (
        inv_qs.annotate(y=TruncYear("created_at"))
        .values("y")
        .annotate(total=Sum("total_amount"))
        .order_by("y")
    )
    yearly_map = {r["y"].year: float(r["total"] or 0) for r in yearly_raw if r["y"]}
    yearly = [{"year": y, "total": yearly_map.get(y, 0)} for y in years_list if not year_filter or y == year_filter]

    # Doughnut: revenue received + expenses + COGS + profit/loss
    # includes EBT
    income_sales = inv_qs.filter(
        created_at__year=doughnut_year,
        created_at__month=doughnut_month,
        status__in=["cash", "bank", "ebt"],
        paid=True
    ).aggregate(t=Sum("total_amount"))["t"] or Decimal("0.00")

    income_debt = pay_qs.filter(
        timestamp__year=doughnut_year,
        timestamp__month=doughnut_month
    ).aggregate(t=Sum("amount"))["t"] or Decimal("0.00")

    doughnut_sales_total = Decimal(income_sales) + Decimal(income_debt)

    doughnut_total_expenses = exp_qs.filter(
        date__year=doughnut_year,
        date__month=doughnut_month
    ).aggregate(t=Sum("amount"))["t"] or Decimal("0.00")

    doughnut_inv_qs = inv_qs.filter(
        created_at__year=doughnut_year,
        created_at__month=doughnut_month
    )
    doughnut_items = InvoiceItem.objects.filter(
        invoice__in=doughnut_inv_qs,
        invoice__invoice_state="active"
    )

    if hasattr(InvoiceItem, "line_cost_cache"):
        doughnut_cogs_total = doughnut_items.aggregate(t=Sum("line_cost_cache"))["t"] or Decimal("0.00")
    elif hasattr(InvoiceItem, "buying_cost"):
        doughnut_cogs_total = doughnut_items.aggregate(
            t=Sum(
                ExpressionWrapper(
                    F("buying_cost") * F("quantity"),
                    output_field=DecimalField(max_digits=18, decimal_places=2)
                )
            )
        )["t"] or Decimal("0.00")
    else:
        doughnut_cogs_total = Decimal("0.00")

    doughnut_net_profit = doughnut_sales_total - doughnut_total_expenses - doughnut_cogs_total
    profit_is_negative = doughnut_net_profit < 0
    doughnut_profit_for_chart = doughnut_net_profit if doughnut_net_profit > 0 else Decimal("0.00")

    month_choices = [
        (1, "January"), (2, "February"), (3, "March"), (4, "April"),
        (5, "May"), (6, "June"), (7, "July"), (8, "August"),
        (9, "September"), (10, "October"), (11, "November"), (12, "December"),
    ]
    year_choices = list(range(today.year - 7, today.year + 1))

    return _render(request, ["sales/enhanced_sales_report.html", "enhanced_sales_report.html"], {
        "branches": _branches_for_sales_dropdown() if _can_filter_branch(request) else [],
        "selected_branch": str(selected.id) if selected else "",
        "selected_branch_obj": selected,

        "daily_sales_json": json.dumps(daily),
        "weekly_sales_json": json.dumps(weekly),
        "monthly_sales_json": json.dumps(monthly),
        "yearly_sales_json": json.dumps(yearly),

        "date_filter": date_filter,
        "month_filter": month_filter or "",
        "year_filter": year_filter or "",

        "month_choices": month_choices,
        "year_choices": year_choices,

        "doughnut_month": doughnut_month,
        "doughnut_year": doughnut_year,
        "doughnut_sales_total": float(doughnut_sales_total),
        "doughnut_total_expenses": float(doughnut_total_expenses),
        "doughnut_cogs_total": float(doughnut_cogs_total),
        "doughnut_net_profit": float(doughnut_net_profit),
        "doughnut_profit_for_chart": float(doughnut_profit_for_chart),
        "profit_is_negative": profit_is_negative,
    })
# ───────────────────────────────────────────────────────────────
# EXPENSES
# ───────────────────────────────────────────────────────────────
@login_required
@role_required(["admin", "cashier", "seller"])
def expense_overview(request):
    today = localdate()

    branch_id = (request.GET.get("branch") or "").strip()
    selected_branch = None

    if _is_admin(request):
        if branch_id:
            selected_branch = Branch.objects.filter(pk=branch_id, is_active=True).first()
    else:
        selected_branch = _get_user_branch(request)

    inv_qs = Invoice.objects.all()
    exp_qs = Expense.objects.all()

    if selected_branch and hasattr(Invoice, "branch_id"):
        inv_qs = inv_qs.filter(branch=selected_branch)
    if selected_branch and hasattr(Expense, "branch_id"):
        exp_qs = exp_qs.filter(branch=selected_branch)

    # ✅ include EBT here
    sales_today = (
        inv_qs.filter(created_at__date=today, status__in=["cash", "bank", "ebt"], paid=True)
        .aggregate(t=Sum("total_amount"))["t"] or Decimal("0.00")
    )
    expenses_today = exp_qs.filter(date=today).aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
    net_profit = sales_today - expenses_today

    today_expenses_list = exp_qs.filter(date=today).order_by("-id")[:20]
    branches = Branch.objects.filter(is_active=True).order_by("name") if _is_admin(request) else None

    return _render(request, ["sales/expense_overview.html", "expense_overview.html"], {
        "sales_today": sales_today,
        "expenses_today": expenses_today,
        "net_profit": net_profit,
        "today_expenses_list": today_expenses_list,

        "branches": branches,
        "selected_branch": str(selected_branch.id) if selected_branch else "",
        "selected_branch_obj": selected_branch,
        "is_admin": _is_admin(request),
    })
# ───────────────────────────────────────────────────────────────
# PRODUCTS (seller sees ONLY branch stock)
# ───────────────────────────────────────────────────────────────
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import OuterRef, Subquery, IntegerField, Value, Q
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


@login_required
@role_required(["admin", "cashier", "seller"])
def product_list(request):
    q = (request.GET.get("q") or "").strip()
    selected_branch_id = (request.GET.get("branch") or "").strip()
    selected_business_type = (request.GET.get("business_type") or "").strip()

    user_branch = _get_user_branch(request) if _is_seller(request) else None
    is_seller = bool(user_branch)

    qs = Product.objects.select_related("category")

    # Branch list for filter dropdown
    branches = Branch.objects.filter(is_active=True).order_by("name") if 'Branch' in globals() else Branch.objects.all().order_by("name")

    selected_branch = None
    use_branch_stock = False

    # Seller locked to own branch
    if is_seller:
        selected_branch = user_branch
        use_branch_stock = True

        ps = ProductStock.objects.filter(
            product=OuterRef("pk"),
            branch=selected_branch
        ).values("quantity")[:1]

        qs = qs.annotate(
            branch_qty=Coalesce(Subquery(ps, output_field=IntegerField()), Value(0))
        ).filter(branch_qty__gt=0)

    else:
        # Admin/Cashier can optionally filter by branch
        if selected_branch_id.isdigit():
            selected_branch = Branch.objects.filter(id=selected_branch_id).first()

        if selected_branch:
            use_branch_stock = True
            ps = ProductStock.objects.filter(
                product=OuterRef("pk"),
                branch=selected_branch
            ).values("quantity")[:1]

            qs = qs.annotate(
                branch_qty=Coalesce(Subquery(ps, output_field=IntegerField()), Value(0))
            ).filter(branch_qty__gt=0)
        else:
            use_branch_stock = False

    # Business type filter
    valid_business_types = {"electronics", "furniture", "magodoro", "unassigned"}
    if selected_business_type in valid_business_types:
        qs = qs.filter(business_type=selected_business_type)
    else:
        selected_business_type = ""

    # Search filter
    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(category__name__icontains=q) |
            Q(business_type__icontains=q)
        )

    paginator = Paginator(qs.order_by("-id"), 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    return _render(request, ["sales/product_list.html", "product_list.html"], {
        "products": page_obj.object_list,
        "is_paginated": page_obj.has_other_pages(),
        "page_obj": page_obj,
        "paginator": paginator,
        "query": q,
        "use_branch_stock": use_branch_stock,
        "branch": selected_branch,
        "branches": branches,
        "selected_branch_id": str(selected_branch.id) if selected_branch else "",
        "selected_business_type": selected_business_type,
        "business_type_choices": [
            ("electronics", "Electronics"),
            ("furniture", "Furniture"),
            ("magodoro", "Magodoro"),
            ("unassigned", "Unassigned"),
        ],
        "is_seller": is_seller,
    })


@login_required
@role_required(["admin", "cashier", "seller"])
def export_products_excel(request):
    q = (request.GET.get("q") or "").strip()
    selected_branch_id = (request.GET.get("branch") or "").strip()

    user_branch = _get_user_branch(request) if _is_seller(request) else None
    is_seller = bool(user_branch)

    qs = Product.objects.select_related("category")

    selected_branch = None
    use_branch_stock = False

    if is_seller:
        selected_branch = user_branch
        use_branch_stock = True

        ps = ProductStock.objects.filter(
            product=OuterRef("pk"),
            branch=selected_branch
        ).values("quantity")[:1]

        qs = qs.annotate(
            branch_qty=Coalesce(Subquery(ps, output_field=IntegerField()), Value(0))
        ).filter(branch_qty__gt=0)

    else:
        if selected_branch_id.isdigit():
            selected_branch = Branch.objects.filter(id=selected_branch_id).first()

        if selected_branch:
            use_branch_stock = True
            ps = ProductStock.objects.filter(
                product=OuterRef("pk"),
                branch=selected_branch
            ).values("quantity")[:1]

            qs = qs.annotate(
                branch_qty=Coalesce(Subquery(ps, output_field=IntegerField()), Value(0))
            ).filter(branch_qty__gt=0)

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(category__name__icontains=q)
        )

    qs = qs.order_by("name")

    wb = Workbook()
    ws = wb.active
    ws.title = "Products Report"

    # ---------- STYLES ----------
    title_font = Font(bold=True, size=16, color="1E3A8A")
    subtitle_font = Font(bold=False, size=10, color="475569")
    header_font = Font(bold=True, color="FFFFFF")
    body_font = Font(size=11, color="0F172A")

    header_fill = PatternFill("solid", fgColor="2563EB")
    thin_side = Side(style="thin", color="CBD5E1")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # ---------- PAGE / PRINT SETUP ----------
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "4:4"

    # ---------- TITLE ----------
    ws.merge_cells("A1:C1")
    ws["A1"] = "MSUMARI JR - PRODUCTS STOCK REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = left

    branch_label = selected_branch.name if selected_branch else "All Branches"
    ws.merge_cells("A2:C2")
    ws["A2"] = f"Branch: {branch_label}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = left

    ws.merge_cells("A3:C3")
    ws["A3"] = f"Search Filter: {q if q else 'All Products'}"
    ws["A3"].font = subtitle_font
    ws["A3"].alignment = left

    # ---------- HEADERS ----------
    headers = ["PRODUCT", "CATEGORY", "CURRENT STOCK"]
    header_row = 4

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # ---------- DATA ----------
    start_row = 5
    row_num = start_row

    for product in qs:
        stock = product.branch_qty if use_branch_stock else (product.stock or 0)

        ws.cell(row=row_num, column=1, value=product.name or "")
        ws.cell(row=row_num, column=2, value=product.category.name if product.category else "-")
        ws.cell(row=row_num, column=3, value=stock)

        ws.cell(row=row_num, column=1).font = body_font
        ws.cell(row=row_num, column=2).font = body_font
        ws.cell(row=row_num, column=3).font = body_font

        ws.cell(row=row_num, column=1).alignment = left
        ws.cell(row=row_num, column=2).alignment = left
        ws.cell(row=row_num, column=3).alignment = center

        ws.cell(row=row_num, column=1).border = border
        ws.cell(row=row_num, column=2).border = border
        ws.cell(row=row_num, column=3).border = border

        row_num += 1

    # ---------- EMPTY CASE ----------
    if row_num == start_row:
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=3)
        ws["A5"] = "No products found."
        ws["A5"].alignment = center
        ws["A5"].font = Font(bold=True, color="64748B")

    # ---------- COLUMN WIDTHS ----------
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18

    # ---------- TOTAL ----------
    total_row = row_num + 1
    ws.cell(row=total_row, column=1, value="Total Products")
    ws.cell(row=total_row, column=1).font = Font(bold=True, color="0F172A")
    ws.cell(row=total_row, column=1).border = border
    ws.cell(row=total_row, column=1).alignment = left

    ws.cell(row=total_row, column=2, value="")
    ws.cell(row=total_row, column=2).border = border

    ws.cell(row=total_row, column=3, value=max(row_num - start_row, 0))
    ws.cell(row=total_row, column=3).font = Font(bold=True, color="0F172A")
    ws.cell(row=total_row, column=3).border = border
    ws.cell(row=total_row, column=3).alignment = center

    filename = "products_report.xlsx"
    if selected_branch:
        safe_branch = "".join(c for c in selected_branch.name if c.isalnum() or c in (" ", "_", "-")).strip().replace(" ", "_")
        filename = f"products_report_{safe_branch}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@role_required(["admin", "cashier", "seller"])
def product_details_json(request, pk):
    p = get_object_or_404(Product.objects.select_related("category"), pk=pk)
    branch = _get_user_branch(request) if _is_seller(request) else None

    entries_qs = StockEntry.objects.filter(product=p)
    if branch:
        entries_qs = entries_qs.filter(branch=branch)

    entries = entries_qs.order_by("timestamp")
    history = [{
        "date": datefmt(e.timestamp, "Y-m-d H:i"),
        "change": e.change,
        "note": e.note,
        "branch": e.branch.name if e.branch else "",
    } for e in entries]

    current_stock = _get_branch_qty(p.pk, branch) if branch else int(getattr(p, "stock", 0) or 0)

    return JsonResponse({
        "name": p.name,
        "category": p.category.name if p.category else "",
        "selling_price": f"{_product_selling_price(p):.2f}",
        "buying_price": f"{_product_buying_price(p):.2f}",
        "current_stock": current_stock,
        "history": history,
    })

@login_required
@role_required(["admin", "cashier", "seller"])
def product_detail(request, pk):
    product = get_object_or_404(Product.objects.select_related("category"), pk=pk)

    is_admin = _is_admin(request)
    branch = _get_user_branch(request) if _is_seller(request) else None

    current_stock = _get_branch_qty(product.pk, branch) if branch else int(getattr(product, "stock", 0) or 0)

    if request.method == "POST":
        if not is_admin:
            return HttpResponseForbidden("Only admin can manage stock.")

        action = (request.POST.get("action") or "add").strip().lower()
        note = (request.POST.get("note") or "").strip()
        branch_id = (request.POST.get("branch_id") or "").strip()

        target_branch = Branch.objects.filter(pk=branch_id).first() if branch_id else None
        if not target_branch:
            target_branch = Branch.objects.filter(is_active=True).first() or Branch.objects.first()

        if not target_branch:
            messages.error(request, "No branch found. Please create a branch first.")
            return redirect("sales:product_detail", pk=pk)

        try:
            with transaction.atomic():
                current_branch_qty = int(
                    ProductStock.objects.filter(product=product, branch=target_branch)
                    .values_list("quantity", flat=True)
                    .first() or 0
                )

                if action == "reset":
                    if current_branch_qty <= 0:
                        messages.warning(
                            request,
                            f"'{product.name}' already has 0 stock in {target_branch.name}."
                        )
                        return redirect("sales:product_detail", pk=pk)

                    StockEntry.objects.create(
                        product=product,
                        change=-current_branch_qty,
                        branch=target_branch,
                        note=note or f"Stock reset to zero in {target_branch.name}",
                    )

                    messages.success(
                        request,
                        f"Stock for '{product.name}' in {target_branch.name} has been reset to 0."
                    )

                else:
                    add_qty = _safe_int(request.POST.get("add_qty"), 0) or 0

                    if add_qty <= 0:
                        messages.error(request, "Enter a valid quantity greater than 0.")
                        return redirect("sales:product_detail", pk=pk)

                    StockEntry.objects.create(
                        product=product,
                        change=add_qty,
                        branch=target_branch,
                        note=note or "Restocked via UI",
                    )

                    messages.success(
                        request,
                        f"Added {add_qty} to '{product.name}' in {target_branch.name}."
                    )

        except (ValidationError, ValueError) as e:
            messages.error(request, str(e))

        return redirect("sales:product_detail", pk=pk)

    entries_qs = StockEntry.objects.filter(product=product)
    if branch:
        entries_qs = entries_qs.filter(branch=branch)
    entries = entries_qs.order_by("-timestamp")

    branch_stocks = []
    if is_admin:
        branch_stocks = (
            ProductStock.objects
            .filter(product=product)
            .select_related("branch")
            .order_by("branch__name")
        )

    current_stock = _get_branch_qty(product.pk, branch) if branch else int(getattr(product, "stock", 0) or 0)

    return _render(request, ["sales/product_detail.html", "product_detail.html"], {
        "product": product,
        "entries": entries,
        "is_admin": is_admin,
        "branch": branch,
        "use_branch_stock": bool(branch),
        "current_stock": current_stock,
        "branches": Branch.objects.filter(is_active=True).order_by("name") if is_admin else [],
        "branch_stocks": branch_stocks,
    })

@login_required
@role_required(["admin", "cashier", "seller"])
def top_selling_products(request):
    q = (request.GET.get("q") or "").strip()
    selected = _selected_branch(request)   # ✅ seller locked, admin/cashier optional

    qs = InvoiceItem.objects.select_related("product", "invoice")

    # ✅ apply branch scope
    if selected and hasattr(Invoice, "branch_id"):
        qs = qs.filter(invoice__branch=selected)
    elif _is_seller(request) and not hasattr(Invoice, "branch_id"):
        # fallback if Invoice has no branch field
        qs = qs.filter(invoice__seller=request.user)

    if q:
        qs = qs.filter(product__name__icontains=q)

    aggregated = (
        qs.values("product_id")
        .annotate(total_sold=Sum("quantity"))
        .order_by("-total_sold", "product_id")
    )

    product_ids = [r["product_id"] for r in aggregated]
    product_map = {
        p.id: p for p in Product.objects.filter(id__in=product_ids)
    }

    items = [
        {
            "product": product_map.get(r["product_id"]),
            "total_sold": r["total_sold"] or 0
        }
        for r in aggregated
        if product_map.get(r["product_id"])
    ]

    paginator = Paginator(items, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    return _render(request, ["sales/top_selling_products.html", "top_selling_products.html"], {
        "top_products": page_obj.object_list,
        "is_paginated": page_obj.has_other_pages(),
        "page_obj": page_obj,
        "paginator": paginator,
        "query": q,

        # ✅ branch filter context
        "branches": _branches_for_sales_dropdown() if _can_filter_branch(request) else [],
        "selected_branch": str(selected.id) if selected else "",
        "selected_branch_obj": selected,

        "is_admin": _is_admin(request),
        "is_cashier": _is_cashier(request),
        "is_seller": _is_seller(request),
    })

# ───────────────────────────────────────────────────────────────
# ADMIN ONLY PAGES
# ───────────────────────────────────────────────────────────────
@login_required
@role_required(["admin"])
def customer_list(request):
    q = (request.GET.get("q") or "").strip()

    rows = _collect_all_contacts()

    if q:
        ql = q.lower()
        rows = [
            r for r in rows
            if ql in (r.get("name", "") or "").lower()
            or ql in (r.get("phone", "") or "").lower()
            or ql in (r.get("source", "") or "").lower()
        ]

    paginator = Paginator(rows, 25)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return _render(request, ["sales/customer_list.html", "customer_list.html"], {
        "customers": page_obj.object_list,
        "page_obj": page_obj,
        "is_paginated": page_obj.has_other_pages(),
        "query": q,
    })


@login_required
@role_required(["admin"])
def supplier_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = Supplier.objects.all().order_by("-id")
    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(phone__icontains=q) |
            Q(contact__icontains=q) |
            Q(address__icontains=q)
        )

    page_obj = Paginator(qs, 25).get_page(request.GET.get("page"))

    return _render(request, ["sales/supplier_list.html", "supplier_list.html"], {
        "suppliers": page_obj.object_list,
        "page_obj": page_obj,
        "is_paginated": page_obj.has_other_pages(),
        "query": q,
    })


@login_required
@role_required(["admin"])
def purchase_invoices(request):
    qs = Invoice.objects.filter(status="debt", paid=False)
    return _render(request, ["sales/invoice_list.html", "invoice_list.html"], {
        "invoices": qs,
        "title": "Purchase Invoices Due",
    })


@login_required
@role_required(["admin"])
def sales_invoices(request):
    qs = Invoice.objects.filter(status="cash")
    return _render(request, ["sales/invoice_list.html", "invoice_list.html"], {
        "invoices": qs,
        "title": "Sales Invoices",
    })


@login_required
@role_required(["admin"])
def all_payments(request):
    qs = Payment.objects.select_related("invoice").order_by("-timestamp")
    return _render(request, ["sales/payment_list.html", "payment_list.html"], {"payments": qs})


# ───────────────────────────────────────────────────────────────
# STOCK MANAGEMENT (ADMIN)
# ───────────────────────────────────────────────────────────────
@login_required
@role_required(["admin"])
def stock_management(request):
    """
    Admin Stock Management:
    - Choose branch (sales branches only)
    - Search product
    - Add (restock) or Reduce stock
    - Stock updates ONLY via StockEntry
    - Shows recent movements
    """
    branches = _branches_for_sales_dropdown()

    branch_id = (request.GET.get("branch") or request.POST.get("branch") or "").strip()
    selected_branch = branches.filter(pk=branch_id).first() if branch_id else branches.first()

    movements_qs = StockEntry.objects.select_related("product", "branch").order_by("-timestamp")
    if selected_branch:
        movements_qs = movements_qs.filter(branch=selected_branch)

    if request.method == "POST":
        product_id = (request.POST.get("product_id") or "").strip()
        qty = _safe_int(request.POST.get("quantity"), 0) or 0
        action = (request.POST.get("action") or "add").strip().lower()
        note = (request.POST.get("note") or "").strip()

        if not selected_branch:
            messages.error(request, "Please select a branch.")
        elif not product_id:
            messages.error(request, "Please choose a product.")
        elif qty <= 0:
            messages.error(request, "Enter a valid quantity (greater than 0).")
        else:
            product = get_object_or_404(Product, pk=product_id)
            if action == "reduce":
                delta = -abs(int(qty))
                action_label = "Reduced"
            else:
                delta = abs(int(qty))
                action_label = "Added"

            try:
                with transaction.atomic():
                    StockEntry.objects.create(
                        product=product,
                        branch=selected_branch,
                        change=delta,
                        note=note or f"{action_label} via Stock Management",
                    )

                    new_qty = (
                        ProductStock.objects.filter(product=product, branch=selected_branch)
                        .values_list("quantity", flat=True)
                        .first()
                    ) or 0

                messages.success(
                    request,
                    f"{action_label} {abs(delta)} to '{product.name}' ({selected_branch.name}). New stock: {new_qty}"
                )
            except Exception as e:
                messages.error(request, str(e))

    return _render(request, ["sales/stock_management.html", "stock_management.html"], {
        "branches": branches,
        "selected_branch": str(selected_branch.id) if selected_branch else "",
        "selected_branch_obj": selected_branch,
        "movements": movements_qs[:60],
        "today_str": localdate().isoformat(),
    })


@login_required
@role_required(["admin"])
def stock_search_products(request):
    """
    AJAX product search for Stock Management:
    /sales/stock/search/?q=...&branch=ID
    Returns: [{id,name,price,stock}]
    """
    q = (request.GET.get("q") or "").strip()
    branch_id = (request.GET.get("branch") or "").strip()

    branch = Branch.objects.filter(pk=branch_id).first() if branch_id else None

    qs = Product.objects.select_related("category").order_by("name")

    if branch:
        ps = ProductStock.objects.filter(product=OuterRef("pk"), branch=branch).values("quantity")[:1]
        qs = qs.annotate(branch_qty=Coalesce(Subquery(ps, output_field=IntegerField()), Value(0)))
    else:
        qs = qs.annotate(branch_qty=Value(0, output_field=IntegerField()))

    if q:
        qs = qs.filter(name__icontains=q)

    results = []
    for p in qs[:30]:
        results.append({
            "id": p.id,
            "name": p.name,
            "price": float(_product_selling_price(p) or 0),
            "stock": int(getattr(p, "branch_qty", 0) or 0),
        })

    return JsonResponse(results, safe=False)
# sales/views.py
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from django.db.models import Sum
from django.core.paginator import Paginator
from django.utils import timezone

from .models import Customer, Invoice, Payment
from users.utils import role_required

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.utils import timezone

from users.utils import role_required
from .models import Customer, Invoice, Payment, SalesReturn  # ✅ add SalesReturn

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.utils import timezone

from users.utils import role_required
from .models import Customer, Invoice, Payment, SalesReturn


@login_required
@role_required(["admin", "cashier", "seller"])
def customer_debt_statement(request, customer_id):
    customer = get_object_or_404(Customer, pk=customer_id)

    invoices_qs = (
        Invoice.objects
        .filter(customer=customer, status="debt")
        .prefetch_related("items__product", "payments", "returns")
        .order_by("-created_at")
    )

    total_debt = invoices_qs.aggregate(t=Sum("total_amount"))["t"] or Decimal("0")

    total_paid_int = Payment.objects.filter(
        invoice__customer=customer,
        invoice__status="debt",
    ).aggregate(t=Sum("amount"))["t"] or 0
    total_paid = Decimal(total_paid_int)

    returns_qs = SalesReturn.objects.filter(
        invoice__customer=customer,
        invoice__status="debt",
    )
    if hasattr(SalesReturn, "status"):
        returns_qs = returns_qs.filter(status__in=["posted", "POSTED", "Posted"])

    total_returns = returns_qs.aggregate(t=Sum("total_amount"))["t"] or Decimal("0")

    # ✅ NET DEBT (this is what you expect to "decrease")
    net_debt = Decimal(total_debt) - Decimal(total_returns)

    # ✅ balance can be negative -> customer credit
    balance = net_debt - Decimal(total_paid)

    credit_due = Decimal("0")
    if balance < 0:
        credit_due = -balance  # shop owes customer
        # keep balance negative to show it properly

    invoices_page = Paginator(invoices_qs, 10).get_page(request.GET.get("page", 1))

    # ---------- Statement events ----------
    invoices_for_statement = (
        Invoice.objects
        .filter(customer=customer, status="debt")
        .prefetch_related("items__product", "payments", "returns")
        .order_by("created_at")
    )

    events = []

    for inv in invoices_for_statement:
        items_count = inv.items.all().count()
        inv_total = Decimal(inv.total_amount or 0)

        # Dr invoice
        events.append({
            "dt": inv.created_at,
            "ref": f"INV#{inv.id}",
            "desc": f"Goods taken ({items_count} items)",
            "debit": inv_total,
            "credit": Decimal("0"),
            "kind": "invoice",
        })

        # Cr payments
        for p in inv.payments.all():
            who = f" by {p.created_by.username}" if getattr(p, "created_by", None) else ""
            events.append({
                "dt": p.timestamp,
                "ref": f"PAY#{p.id}",
                "desc": f"Payment{who}",
                "debit": Decimal("0"),
                "credit": Decimal(p.amount or 0),
                "kind": "payment",
            })

        # Cr returns
        inv_returns = inv.returns.all()
        if hasattr(SalesReturn, "status"):
            inv_returns = inv_returns.filter(status__in=["posted", "POSTED", "Posted"])

        for r in inv_returns:
            who = f" by {r.created_by.username}" if getattr(r, "created_by", None) else ""
            events.append({
                "dt": getattr(r, "created_at", None) or timezone.now(),
                "ref": f"RET#{r.id}",
                "desc": f"Return{who}",
                "debit": Decimal("0"),
                "credit": Decimal(getattr(r, "total_amount", 0) or 0),
                "kind": "return",
            })

    events.sort(key=lambda x: x["dt"] or timezone.now())

    # ✅ running balance (DON'T clamp to 0)
    running = Decimal("0")
    rows = []
    for e in events:
        running = running + Decimal(e.get("debit") or 0) - Decimal(e.get("credit") or 0)
        e["balance"] = running
        rows.append(e)

    return render(request, "sales/customer_debt_statement.html", {
        "customer": customer,
        "invoices": invoices_page,

        "total_debt": total_debt,
        "total_paid": total_paid,
        "total_returns": total_returns,

        # ✅ NEW
        "net_debt": net_debt,          # Debt - Returns
        "balance": balance,            # can be negative
        "credit_due": credit_due,      # positive number if customer overpaid

        "rows": rows,
    })


# sales/views.py (stock transfer part)

import json
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import JsonResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render

from inventory.models import Branch, Product, ProductStock
from users.utils import role_required
from .models import StockTransfer
from .forms import StockTransferForm, StockTransferItemFormSet


def sales_branches_qs():
    return (
        Branch.objects.filter(is_active=True)
        .exclude(name__icontains="transport")
        .exclude(name__icontains="hama na")
        .exclude(name__icontains="hq")
        .order_by("name")
    )


def user_branch(request):
    prof = getattr(request.user, "profile", None)
    return getattr(prof, "branch", None)


def user_role(request):
    prof = getattr(request.user, "profile", None)
    return (getattr(prof, "role", "") or "").lower()


@login_required
@role_required(["admin", "seller", "cashier"])
def stock_transfer_list(request):
    role = user_role(request)
    branch = user_branch(request)

    qs = StockTransfer.objects.select_related("from_branch", "to_branch", "created_by", "posted_by")

    # seller sees only their transfers
    if role == "seller" and branch:
        qs = qs.filter(Q(from_branch=branch) | Q(to_branch=branch))

    # branch filter dropdown (✅ sales branches only)
    branches = sales_branches_qs()

    f_branch = (request.GET.get("branch") or "").strip()
    if f_branch:
        qs = qs.filter(Q(from_branch_id=f_branch) | Q(to_branch_id=f_branch))

    status = (request.GET.get("status") or "").strip()
    if status:
        qs = qs.filter(status=status)

    transfers = qs.order_by("-id")[:500]

    return render(request, "sales/stock_transfer_list.html", {
        "transfers": transfers,
        "branches": branches,
        "selected_branch": f_branch,
        "selected_status": status,
    })


@login_required
@role_required(["admin", "seller"])
def stock_transfer_create(request):
    role = user_role(request)
    branch = user_branch(request)

    if request.method == "POST":
        form = StockTransferForm(request.POST, user=request.user)
        formset = StockTransferItemFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            transfer = form.save(commit=False)
            transfer.created_by = request.user

            # ✅ seller: force from_branch
            if role == "seller" and branch:
                transfer.from_branch = branch

            transfer.save()
            formset.instance = transfer
            formset.save()

            messages.success(request, f"Transfer #{transfer.id} created.")
            return redirect("sales:stock_transfer_detail", transfer.id)

        messages.error(request, "Please fix the errors below.")
    else:
        form = StockTransferForm(user=request.user)
        formset = StockTransferItemFormSet()

    products_json = json.dumps(list(Product.objects.values("id", "name")))

    return render(request, "sales/stock_transfer_form.html", {
        "form": form,
        "formset": formset,
        "products_json": products_json,
    })


@login_required
@role_required(["admin", "seller", "cashier"])
def stock_transfer_detail(request, pk):
    transfer = get_object_or_404(
        StockTransfer.objects.select_related("from_branch", "to_branch", "created_by", "posted_by"),
        pk=pk
    )

    role = user_role(request)
    branch = user_branch(request)

    # ✅ seller permission
    if role == "seller" and branch:
        if transfer.from_branch_id != branch.id and transfer.to_branch_id != branch.id:
            return HttpResponseForbidden("You don't have permission.")

    if request.method == "POST" and transfer.status == StockTransfer.STATUS_DRAFT and request.POST.get("save_items") == "1":
        form = StockTransferForm(request.POST, instance=transfer, user=request.user)
        formset = StockTransferItemFormSet(request.POST, instance=transfer)

        if form.is_valid() and formset.is_valid():
            t = form.save(commit=False)

            # ✅ seller keep locked
            if role == "seller" and branch:
                t.from_branch = branch

            t.save()
            formset.save()

            messages.success(request, "Transfer updated.")
            return redirect("sales:stock_transfer_detail", transfer.id)
        else:
            messages.error(request, "Fix errors below.")
    else:
        form = StockTransferForm(instance=transfer, user=request.user)
        formset = StockTransferItemFormSet(instance=transfer)

    return render(request, "sales/stock_transfer_detail.html", {
        "transfer": transfer,
        "form": form,
        "formset": formset,
        "items": transfer.items.select_related("product"),
    })


@login_required
@role_required(["admin", "seller", ])
def stock_transfer_post(request, pk):
    transfer = get_object_or_404(StockTransfer, pk=pk)

    role = user_role(request)
    branch = user_branch(request)

    if request.method != "POST":
        return redirect("sales:stock_transfer_detail", transfer.id)

    # ✅ seller can post only if from_branch == their branch
    if role == "seller" and branch and transfer.from_branch_id != branch.id:
        messages.error(request, "You can only post transfers from your branch.")
        return redirect("sales:stock_transfer_detail", transfer.id)

    try:
        transfer.post(request.user)
        messages.success(request, f"Transfer #{transfer.id} posted successfully.")
    except Exception as e:
        messages.error(request, f"Failed to post transfer: {e}")

    return redirect("sales:stock_transfer_detail", transfer.id)


@login_required
@role_required(["admin", "seller"])
def api_branch_stock(request):
    try:
        branch_id = int(request.GET.get("branch") or 0)
        product_id = int(request.GET.get("product") or 0)
    except ValueError:
        return JsonResponse({"ok": False, "qty": 0, "message": "Invalid params"}, status=400)

    if not branch_id or not product_id:
        return JsonResponse({"ok": True, "qty": 0})

    ps = ProductStock.objects.filter(branch_id=branch_id, product_id=product_id).first()
    qty = int(getattr(ps, "quantity", 0) or 0)
    return JsonResponse({"ok": True, "qty": qty})



# sales/views.py
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from users.utils import role_required
from .models import Debtor
from .forms import DebtorForm

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum
from django.shortcuts import render
from users.utils import role_required

from .models import Debtor, DebtorPayment  # adjust if your payment model name differs


@login_required
@role_required(["admin"])
def debtor_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = Debtor.objects.all().order_by("-id")
    if q:
        qs = qs.filter(name__icontains=q)

    # Build rows with computed totals to avoid template errors
    rows = []
    total_owed_active = Decimal("0.00")
    active_count = 0
    closed_count = 0

    for d in qs:
        paid = DebtorPayment.objects.filter(debtor=d).aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
        owed = Decimal(str(d.amount_owed or 0))
        balance = owed - Decimal(str(paid or 0))
        if balance < 0:
            balance = Decimal("0.00")

        # ✅ auto close if cleared
        is_active = bool(balance > 0)
        if d.is_active != is_active:
            d.is_active = is_active
            d.save(update_fields=["is_active"])

        if is_active:
            total_owed_active += balance
            active_count += 1
        else:
            closed_count += 1

        rows.append({
            "obj": d,
            "paid": paid,
            "balance": balance,
        })

    # paginate rows
    paginator = Paginator(rows, 20)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(request, "sales/debtor_list.html", {
        "debtors": page_obj.object_list,   # list of dicts: {obj, paid, balance}
        "page_obj": page_obj,
        "query": q,

        # ✅ names your template needs
        "total_owed": total_owed_active,
        "active_count": active_count,
        "closed_count": closed_count,
    })


@login_required
@role_required(["admin"])
def debtor_create(request):
    if request.method == "POST":
        form = DebtorForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, "Debtor added successfully.")
            return redirect("sales:debtor_list")
        messages.error(request, "Please fix the errors below.")
    else:
        form = DebtorForm()

    return render(request, "sales/debtor_form.html", {
        "form": form,
        "is_edit": False,
    })


@login_required
@role_required(["admin"])
def debtor_update(request, pk):
    debtor = get_object_or_404(Debtor, pk=pk)

    if request.method == "POST":
        form = DebtorForm(request.POST, instance=debtor)
        if form.is_valid():
            form.save()
            messages.success(request, "Debtor updated.")
            return redirect("sales:debtor_list")
        messages.error(request, "Please fix the errors below.")
    else:
        form = DebtorForm(instance=debtor)

    return render(request, "sales/debtor_form.html", {
        "form": form,
        "is_edit": True,
        "debtor": debtor,
    })


@login_required
@role_required(["admin"])
def debtor_delete(request, pk):
    debtor = get_object_or_404(Debtor, pk=pk)

    if request.method == "POST":
        debtor.delete()
        messages.success(request, "Debtor deleted.")
        return redirect("sales:debtor_list")

    return render(request, "sales/debtor_delete.html", {
        "debtor": debtor
    })


# sales/views.py
from .models import Debtor, DebtorPayment
from .forms import DebtorForm, DebtorPaymentForm

@login_required
@role_required(["admin"])
def debtor_detail(request, pk):
    debtor = get_object_or_404(Debtor, pk=pk)
    payments = debtor.payments.select_related("created_by").all()

    # payment form
    pay_form = DebtorPaymentForm()

    return render(request, "sales/debtor_detail.html", {
        "debtor": debtor,
        "payments": payments,
        "total_paid": debtor.total_paid(),
        "balance": debtor.balance(),
        "pay_form": pay_form,
    })


@login_required
@role_required(["admin"])
def debtor_add_payment(request, pk):
    debtor = get_object_or_404(Debtor, pk=pk)

    if request.method != "POST":
        return redirect("sales:debtor_detail", pk=debtor.id)

    form = DebtorPaymentForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Fix payment errors.")
        return redirect("sales:debtor_detail", pk=debtor.id)

    amount = Decimal(form.cleaned_data["amount"] or 0)
    if amount <= 0:
        messages.error(request, "Payment must be greater than 0.")
        return redirect("sales:debtor_detail", pk=debtor.id)

    # Don’t allow overpay beyond balance (optional rule)
    current_balance = debtor.balance()
    if amount > current_balance:
        amount = current_balance
        messages.warning(request, f"Payment reduced to balance: {current_balance}")

    try:
        with transaction.atomic():
            p = form.save(commit=False)
            p.debtor = debtor
            p.amount = amount
            p.created_by = request.user
            p.save()

            # ✅ Auto close if finished
            if debtor.balance() <= 0:
                debtor.is_active = False
                debtor.save(update_fields=["is_active", "updated_at"])

    except Exception as e:
        messages.error(request, f"Failed: {e}")
        return redirect("sales:debtor_detail", pk=debtor.id)

    messages.success(request, "Payment recorded.")
    return redirect("sales:debtor_detail", pk=debtor.id)


@login_required
@role_required(["admin"])
def debtor_delete_payment(request, pk, payment_id):
    debtor = get_object_or_404(Debtor, pk=pk)
    payment = get_object_or_404(DebtorPayment, pk=payment_id, debtor=debtor)

    if request.method == "POST":
        payment.delete()

        # ✅ Re-open if balance becomes > 0 again
        if debtor.balance() > 0:
            debtor.is_active = True
            debtor.save(update_fields=["is_active", "updated_at"])

        messages.success(request, "Payment deleted.")
    return redirect("sales:debtor_detail", pk=debtor.id)



# sales/views.py
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from users.utils import role_required

from .models import Invoice, InvoiceItem, SalesReturn, SalesReturnItem
from .forms import SalesReturnForm, SalesReturnItemFormSet


@login_required
@role_required(["admin", "cashier", "seller"])
def sales_return_create(request, invoice_id):
    invoice = get_object_or_404(Invoice, pk=invoice_id)

    # (optional) permissions: seller only own invoice
    # if you have _must_own_invoice_if_seller(request, invoice): use it.

    # Draft return
    ret = SalesReturn(invoice=invoice, created_by=request.user)

    if request.method == "POST":
        form = SalesReturnForm(request.POST, instance=ret)
        formset = SalesReturnItemFormSet(request.POST, instance=ret)

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                ret = form.save(commit=False)
                # set branch from invoice if exists
                inv_branch = getattr(invoice, "branch", None)
                if inv_branch:
                    ret.branch = inv_branch
                ret.created_by = request.user
                ret.save()

                # Save items first
                items = formset.save(commit=False)

                # Only allow returning products that were on invoice
                invoice_products = set(
                    InvoiceItem.objects.filter(invoice=invoice).values_list("product_id", flat=True)
                )

                for it in items:
                    if it.product_id not in invoice_products:
                        raise ValueError("You can only return products from this invoice.")

                    # Pull pricing from invoice item (snapshot)
                    inv_item = InvoiceItem.objects.filter(invoice=invoice, product_id=it.product_id).first()

                    # unit price
                    unit_price = getattr(inv_item, "selling_price", None)
                    if unit_price is None:
                        unit_price = getattr(inv_item, "unit_price", None) or Decimal("0.00")

                    disc = getattr(inv_item, "discount", 0) or 0

                    it.sales_return = ret
                    it.unit_price = Decimal(unit_price)
                    it.discount = Decimal(disc)
                    it.save()

                # delete removed
                for obj in formset.deleted_objects:
                    obj.delete()

                ret.recompute_total()

            messages.success(request, "Return saved as Draft. Now you can POST it.")
            return redirect("sales:sales_return_detail", ret.id)

        messages.error(request, "Please fix the errors below.")
    else:
        form = SalesReturnForm(instance=ret)
        formset = SalesReturnItemFormSet(instance=ret)

    return render(request, "sales/sales_return_form.html", {
        "invoice": invoice,
        "form": form,
        "formset": formset,
    })


@login_required
@role_required(["admin", "cashier", "seller"])
def sales_return_detail(request, pk):
    ret = get_object_or_404(SalesReturn, pk=pk)
    invoice = ret.invoice
    items = ret.items.select_related("product").all()

    return render(request, "sales/sales_return_detail.html", {
        "ret": ret,
        "invoice": invoice,
        "items": items,
    })


@login_required
@role_required(["admin", "cashier", "seller"])
def sales_return_post(request, pk):
    ret = get_object_or_404(SalesReturn, pk=pk)

    if request.method != "POST":
        return redirect("sales:sales_return_detail", ret.id)

    try:
        ret.post(request.user)
        messages.success(request, "Return POSTED. Stock updated. Invoice recalculated.")
    except Exception as e:
        messages.error(request, f"Failed to post return: {e}")

    return redirect("sales:sales_return_detail", ret.id)


@login_required
@role_required(["admin", "cashier", "seller"])
def sales_return_list(request):
    """
    List of returns:
    - Admin/Cashier see all (sales branches)
    - Seller sees only returns from their branch or their invoices
    """
    qs = SalesReturn.objects.select_related("invoice", "created_by")

    # filter by branch if your SalesReturn has branch
    role = _role_lower(request)
    if role == "seller":
        b = _get_user_branch(request)
        if b and hasattr(SalesReturn, "branch_id"):
            qs = qs.filter(branch=b)
        else:
            # fallback: seller sees only their invoices' returns
            qs = qs.filter(invoice__seller=request.user)

    # optional search
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(invoice__id__icontains=q) |
            Q(invoice__customer__name__icontains=q)
        )

    page_obj = Paginator(qs.order_by("-id"), 30).get_page(request.GET.get("page", 1))

    return render(request, "sales/sales_return_list.html", {
        "returns": page_obj.object_list,
        "page_obj": page_obj,
        "is_paginated": page_obj.has_other_pages(),
        "paginator": page_obj.paginator,
        "query": q,
    })





@login_required
@role_required(["admin", "cashier", "seller"])
def period_progress_report(request):
    from decimal import Decimal
    from django.http import HttpResponse
    from django.core.paginator import Paginator
    from django.db.models import Sum
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from .models import SalesReturn

    start_date, end_date, start_str, end_str = _parse_date_range(request)
    selected = _selected_branch(request)

    inv_qs = _invoice_scope_qs(request)
    exp_qs = _expense_scope_qs(request)
    pay_qs = _payment_scope_qs(request)

    if selected and _can_filter_branch(request):
        if hasattr(Invoice, "branch_id"):
            inv_qs = inv_qs.filter(branch=selected)
            pay_qs = pay_qs.filter(invoice__branch=selected)
        if hasattr(Expense, "branch_id"):
            exp_qs = exp_qs.filter(branch=selected)

    inv_period = inv_qs.filter(created_at__date__range=[start_date, end_date]).order_by("-created_at")
    exp_period = exp_qs.filter(date__range=[start_date, end_date]).order_by("-date", "-id")
    pay_period = pay_qs.filter(timestamp__date__range=[start_date, end_date]).order_by("-timestamp")

    ret_qs = SalesReturn.objects.all()
    if hasattr(SalesReturn, "status"):
        ret_qs = ret_qs.filter(status__in=["posted", "POSTED", "Posted"])
    if hasattr(SalesReturn, "created_at"):
        ret_qs = ret_qs.filter(created_at__date__range=[start_date, end_date])

    if _is_seller(request):
        b = _get_user_branch(request)
        if b:
            if hasattr(SalesReturn, "branch_id"):
                ret_qs = ret_qs.filter(branch=b)
            elif hasattr(Invoice, "branch_id"):
                ret_qs = ret_qs.filter(invoice__branch=b)
            else:
                ret_qs = ret_qs.filter(invoice__seller=request.user)
    elif selected:
        if hasattr(SalesReturn, "branch_id"):
            ret_qs = ret_qs.filter(branch=selected)
        elif hasattr(Invoice, "branch_id"):
            ret_qs = ret_qs.filter(invoice__branch=selected)

    cash_sales = _dec(inv_period.filter(status="cash", paid=True).aggregate(t=Sum("total_amount"))["t"] or 0)
    bank_sales = _dec(inv_period.filter(status="bank", paid=True).aggregate(t=Sum("total_amount"))["t"] or 0)
    ebt_sales = _dec(inv_period.filter(status="ebt", paid=True).aggregate(t=Sum("total_amount"))["t"] or 0)
    debt_issued = _dec(inv_period.filter(status="debt").aggregate(t=Sum("total_amount"))["t"] or 0)
    debt_payments_received = _dec(pay_period.aggregate(t=Sum("amount"))["t"] or 0)
    returns_total = _dec(ret_qs.aggregate(t=Sum("total_amount"))["t"] or 0)
    expenses_total = _dec(exp_period.aggregate(t=Sum("amount"))["t"] or 0)

    total_received = cash_sales + bank_sales + ebt_sales + debt_payments_received
    net_cashflow = total_received - expenses_total

    invoice_count = inv_period.count()
    paid_invoice_count = inv_period.filter(status__in=["cash", "bank", "ebt"], paid=True).count()
    debt_invoice_count = inv_period.filter(status="debt").count()

    avg_invoice_value = Decimal("0.00")
    total_invoiced_amount = _dec(inv_period.aggregate(t=Sum("total_amount"))["t"] or 0)
    if invoice_count > 0:
        avg_invoice_value = total_invoiced_amount / Decimal(invoice_count)

    if selected:
        branch_label = selected.name
    elif _is_seller(request):
        seller_branch = _get_user_branch(request)
        branch_label = seller_branch.name if seller_branch else "My Branch"
    else:
        branch_label = "All Branches"

    sales_page = Paginator(inv_period, 15).get_page(request.GET.get("sales_page", 1))
    expense_page = Paginator(exp_period, 15).get_page(request.GET.get("expense_page", 1))

    export_fmt = (request.GET.get("format") or "").strip().lower()
    if export_fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # ---------- Styles ----------
        dark_fill = PatternFill("solid", fgColor="0F172A")
        slate_fill = PatternFill("solid", fgColor="E2E8F0")
        blue_fill = PatternFill("solid", fgColor="E0F2FE")
        green_fill = PatternFill("solid", fgColor="DCFCE7")
        amber_fill = PatternFill("solid", fgColor="FEF3C7")
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        zebra_fill = PatternFill("solid", fgColor="F8FAFC")

        white_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(size=16, bold=True, color="0F172A")
        sub_font = Font(size=11, bold=True, color="475569")
        head_font = Font(bold=True, color="0F172A")
        bold_font = Font(bold=True)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        right = Alignment(horizontal="right", vertical="center")

        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        money_fmt = '#,##0.00'
        int_fmt = '#,##0'

        # ---------- Summary Sheet ----------
        ws.merge_cells("A1:D1")
        ws["A1"] = "MSUMARI JR - SALES PROGRESS REPORT"
        ws["A1"].font = title_font
        ws["A1"].alignment = center

        ws.merge_cells("A2:D2")
        ws["A2"] = f"Period: {start_str} to {end_str} | Branch: {branch_label}"
        ws["A2"].font = sub_font
        ws["A2"].alignment = center

        ws["A4"] = "Metric"
        ws["B4"] = "Value"
        ws["A4"].fill = dark_fill
        ws["B4"].fill = dark_fill
        ws["A4"].font = white_font
        ws["B4"].font = white_font
        ws["A4"].alignment = center
        ws["B4"].alignment = center
        ws["A4"].border = border
        ws["B4"].border = border

        summary_rows = [
            ("Cash Sales", cash_sales, blue_fill),
            ("Bank Sales", bank_sales, blue_fill),
            ("EBT Sales", ebt_sales, blue_fill),
            ("Debt Issued", debt_issued, amber_fill),
            ("Debt Payments Received", debt_payments_received, green_fill),
            ("Returns", returns_total, amber_fill),
            ("Expenses", expenses_total, red_fill),
            ("Total Received", total_received, blue_fill),
            ("Net Cashflow", net_cashflow, green_fill if net_cashflow >= 0 else red_fill),
            ("Invoices Count", invoice_count, slate_fill),
            ("Paid Invoices", paid_invoice_count, green_fill),
            ("Debt Invoices", debt_invoice_count, amber_fill),
            ("Average Invoice", avg_invoice_value, blue_fill),
        ]

        row = 5
        for idx, (label, value, fill) in enumerate(summary_rows, start=1):
            ws[f"A{row}"] = label
            ws[f"B{row}"] = float(value) if isinstance(value, Decimal) else value

            ws[f"A{row}"].font = bold_font
            ws[f"A{row}"].alignment = left
            ws[f"B{row}"].alignment = right

            ws[f"A{row}"].fill = fill
            ws[f"B{row}"].fill = fill

            ws[f"A{row}"].border = border
            ws[f"B{row}"].border = border

            if label in ["Invoices Count", "Paid Invoices", "Debt Invoices"]:
                ws[f"B{row}"].number_format = int_fmt
            else:
                ws[f"B{row}"].number_format = money_fmt

            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.freeze_panes = "A5"

        # ---------- Sales History Sheet ----------
        ws2 = wb.create_sheet("Sales History")
        sales_headers = ["Date", "Invoice", "Customer", "Status", "Paid", "Seller", "Branch", "Total"]

        for c, h in enumerate(sales_headers, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.fill = dark_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        r = 2
        for i, inv in enumerate(inv_period, start=1):
            row_values = [
                getattr(inv, "created_at", None).strftime("%Y-%m-%d %H:%M") if getattr(inv, "created_at", None) else "",
                f"INV-{inv.id}",
                getattr(getattr(inv, "customer", None), "name", "") or "Walk-in",
                (inv.status or "").upper(),
                "YES" if getattr(inv, "paid", False) else "NO",
                getattr(getattr(inv, "seller", None), "username", "") or "",
                getattr(getattr(inv, "branch", None), "name", "") if hasattr(inv, "branch") else "",
                float(_dec(getattr(inv, "total_amount", 0) or 0)),
            ]

            for c, value in enumerate(row_values, 1):
                cell = ws2.cell(r, c, value)
                cell.border = border
                if i % 2 == 0:
                    cell.fill = zebra_fill
                if c == 8:
                    cell.number_format = money_fmt
                    cell.alignment = right
                elif c in [4, 5]:
                    cell.alignment = center

            r += 1

        for col, width in {"A":18, "B":12, "C":24, "D":14, "E":10, "F":16, "G":18, "H":16}.items():
            ws2.column_dimensions[col].width = width

        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:H{max(r-1,1)}"

        # ---------- Expense History Sheet ----------
        ws3 = wb.create_sheet("Expense History")
        expense_headers = ["Date", "Category", "Description", "Branch", "Created By", "Amount"]

        for c, h in enumerate(expense_headers, 1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.fill = dark_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        r = 2
        for i, exp in enumerate(exp_period, start=1):
            row_values = [
                str(getattr(exp, "date", "") or ""),
                getattr(exp, "category", "") or "",
                getattr(exp, "description", "") or "",
                getattr(getattr(exp, "branch", None), "name", "") if hasattr(exp, "branch") else "",
                getattr(getattr(exp, "created_by", None), "username", "") or "",
                float(_dec(getattr(exp, "amount", 0) or 0)),
            ]

            for c, value in enumerate(row_values, 1):
                cell = ws3.cell(r, c, value)
                cell.border = border
                if i % 2 == 0:
                    cell.fill = zebra_fill
                if c == 6:
                    cell.number_format = money_fmt
                    cell.alignment = right

            r += 1

        for col, width in {"A":14, "B":20, "C":34, "D":18, "E":18, "F":16}.items():
            ws3.column_dimensions[col].width = width

        ws3.freeze_panes = "A2"
        ws3.auto_filter.ref = f"A1:F{max(r-1,1)}"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="progress_report_{start_str}_to_{end_str}.xlsx"'
        wb.save(response)
        return response

    return _render(
        request,
        ["sales/period_progress_report.html", "period_progress_report.html"],
        {
            "start_date": start_str,
            "end_date": end_str,

            "branches": _branches_for_sales_dropdown() if _can_filter_branch(request) else [],
            "selected_branch": str(selected.id) if selected else "",
            "selected_branch_obj": selected,
            "branch_label": branch_label,

            "cash_sales": cash_sales,
            "bank_sales": bank_sales,
            "ebt_sales": ebt_sales,
            "debt_issued": debt_issued,
            "debt_payments_received": debt_payments_received,
            "returns_total": returns_total,
            "expenses_total": expenses_total,
            "total_received": total_received,
            "net_cashflow": net_cashflow,

            "invoice_count": invoice_count,
            "paid_invoice_count": paid_invoice_count,
            "debt_invoice_count": debt_invoice_count,
            "avg_invoice_value": avg_invoice_value,

            "sales_page": sales_page,
            "expense_page": expense_page,

            "is_admin": _is_admin(request),
            "is_cashier": _is_cashier(request),
            "is_seller": _is_seller(request),
        }
    )
    
@login_required
@role_required(["admin", "cashier", "seller"])
def genji_dashboard(request):
    start_date, end_date, start_str, end_str = _parse_date_range(request)
    selected = _selected_branch(request)

    qs = GenjiSale.objects.select_related("branch", "created_by")

    if _is_seller(request):
        user_branch = _get_user_branch(request)
        if user_branch:
            qs = qs.filter(branch=user_branch)
        else:
            qs = qs.filter(created_by=request.user)
    elif selected and _can_filter_branch(request):
        qs = qs.filter(branch=selected)

    qs = qs.filter(sell_date__range=[start_date, end_date])

    total_buying = Decimal("0.00")
    total_selling = Decimal("0.00")
    total_profit = Decimal("0.00")

    for row in qs:
        total_buying += row.total_buying
        total_selling += row.total_selling
        total_profit += row.profit

    page_obj = Paginator(qs, 20).get_page(request.GET.get("page", 1))

    return _render(request, ["sales/genji_dashboard.html", "genji_dashboard.html"], {
        "genji_rows": page_obj.object_list,
        "page_obj": page_obj,
        "is_paginated": page_obj.has_other_pages(),
        "paginator": page_obj.paginator,

        "total_buying": total_buying,
        "total_selling": total_selling,
        "total_profit": total_profit,
        "total_records": qs.count(),

        "start_date": start_str,
        "end_date": end_str,

        "branches": _branches_for_sales_dropdown() if _can_filter_branch(request) else [],
        "selected_branch": str(selected.id) if selected else "",
        "selected_branch_obj": selected,

        "is_admin": _is_admin(request),
        "is_cashier": _is_cashier(request),
        "is_seller": _is_seller(request),
    })


@login_required
@role_required(["admin", "cashier", "seller"])
def genji_create(request):
    if request.method == "POST":
        form = GenjiSaleForm(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)

            if _is_seller(request):
                obj.branch = _get_user_branch(request)

            # normalize customer name
            if hasattr(obj, "customer_name"):
                obj.customer_name = _clean_contact_name(
                    getattr(obj, "customer_name", "") or ""
                )

            obj.created_by = request.user
            obj.save()

            # save to customer table too
            _save_customer_contact(
                _first_nonempty_attr(obj, ["customer_name", "name", "client_name"]),
                _first_nonempty_attr(obj, ["customer_phone", "phone", "phone_number", "mobile"]),
            )

            messages.success(request, "Genji record added successfully.")
            return redirect("sales:genji_dashboard")

        messages.error(request, "Please fix the errors below.")
    else:
        form = GenjiSaleForm()

        if _is_seller(request):
            user_branch = _get_user_branch(request)
            if user_branch:
                form.fields["branch"].initial = user_branch
                form.fields["branch"].widget.attrs["disabled"] = True

    return _render(request, ["sales/genji_form.html", "genji_form.html"], {
        "form": form,
        "is_edit": False,
        "today_str": localdate().isoformat(),
        "is_seller": _is_seller(request),
    })

@login_required
@role_required(["admin", "cashier", "seller"])
def genji_update(request, pk):
    obj = get_object_or_404(GenjiSale, pk=pk)

    if _is_seller(request):
        user_branch = _get_user_branch(request)
        if user_branch and obj.branch_id != user_branch.id:
            return HttpResponseForbidden("You cannot edit this record.")

    if request.method == "POST":
        form = GenjiSaleForm(request.POST, instance=obj)
        if form.is_valid():
            updated = form.save(commit=False)

            if _is_seller(request):
                updated.branch = _get_user_branch(request)

            if hasattr(updated, "customer_name"):
                updated.customer_name = _clean_contact_name(
                    getattr(updated, "customer_name", "") or ""
                )

            updated.save()

            _save_customer_contact(
                _first_nonempty_attr(updated, ["customer_name", "name", "client_name"]),
                _first_nonempty_attr(updated, ["customer_phone", "phone", "phone_number", "mobile"]),
            )

            messages.success(request, "Genji record updated.")
            return redirect("sales:genji_dashboard")

        messages.error(request, "Please fix the errors below.")
    else:
        form = GenjiSaleForm(instance=obj)
        if _is_seller(request):
            form.fields["branch"].widget.attrs["disabled"] = True

    return _render(request, ["sales/genji_form.html", "genji_form.html"], {
        "form": form,
        "is_edit": True,
        "obj": obj,
        "is_seller": _is_seller(request),
    })


@login_required
@role_required(["admin", "cashier", "seller"])
def genji_delete(request, pk):
    obj = get_object_or_404(GenjiSale, pk=pk)

    if _is_seller(request):
        user_branch = _get_user_branch(request)
        if user_branch and obj.branch_id != user_branch.id:
            return HttpResponseForbidden("You cannot delete this record.")

    if request.method == "POST":
        obj.delete()
        messages.success(request, "Genji record deleted.")
        return redirect("sales:genji_dashboard")

    return _render(request, ["sales/genji_delete.html", "genji_delete.html"], {
        "obj": obj,
    })
    
@login_required
@role_required(["admin", "cashier"])
def cancel_invoice(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")

    try:
        invoice.cancel_invoice(user=request.user)
        messages.success(request, f"Invoice #{invoice.id} cancelled successfully. Stock restored.")
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Failed to cancel invoice: {e}")

    return redirect("sales:invoice_detail", pk=invoice.pk)


@login_required
@role_required(["admin"])
def my_debt_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = MyDebt.objects.all().order_by("-id")
    if q:
        qs = qs.filter(name__icontains=q)

    rows = []
    total_owed_active = Decimal("0.00")
    active_count = 0
    closed_count = 0

    for d in qs:
        paid = MyDebtPayment.objects.filter(debt=d).aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
        owed = Decimal(str(d.amount_owed or 0))
        balance = owed - Decimal(str(paid or 0))
        if balance < 0:
            balance = Decimal("0.00")

        is_active = bool(balance > 0)
        if d.is_active != is_active:
            d.is_active = is_active
            d.save(update_fields=["is_active"])

        if is_active:
            total_owed_active += balance
            active_count += 1
        else:
            closed_count += 1

        rows.append({
            "obj": d,
            "paid": paid,
            "balance": balance,
        })

    paginator = Paginator(rows, 20)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(request, "sales/my_debt_list.html", {
        "debts": page_obj.object_list,
        "page_obj": page_obj,
        "query": q,
        "total_owed": total_owed_active,
        "active_count": active_count,
        "closed_count": closed_count,
    })


@login_required
@role_required(["admin"])
def my_debt_create(request):
    if request.method == "POST":
        form = MyDebtForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, "Debt added successfully.")
            return redirect("sales:my_debt_list")
        messages.error(request, "Please fix the errors below.")
    else:
        form = MyDebtForm()

    return render(request, "sales/my_debt_form.html", {
        "form": form,
        "is_edit": False,
    })


@login_required
@role_required(["admin"])
def my_debt_update(request, pk):
    debt = get_object_or_404(MyDebt, pk=pk)

    if request.method == "POST":
        form = MyDebtForm(request.POST, instance=debt)
        if form.is_valid():
            form.save()
            messages.success(request, "Debt updated.")
            return redirect("sales:my_debt_list")
        messages.error(request, "Please fix the errors below.")
    else:
        form = MyDebtForm(instance=debt)

    return render(request, "sales/my_debt_form.html", {
        "form": form,
        "is_edit": True,
        "debt": debt,
    })


@login_required
@role_required(["admin"])
def my_debt_delete(request, pk):
    debt = get_object_or_404(MyDebt, pk=pk)

    if request.method == "POST":
        debt.delete()
        messages.success(request, "Debt deleted.")
        return redirect("sales:my_debt_list")

    return render(request, "sales/my_debt_delete.html", {
        "debt": debt
    })


@login_required
@role_required(["admin"])
def my_debt_detail(request, pk):
    debt = get_object_or_404(MyDebt, pk=pk)
    payments = debt.payments.select_related("created_by").all()
    pay_form = MyDebtPaymentForm()

    return render(request, "sales/my_debt_detail.html", {
        "debt": debt,
        "payments": payments,
        "total_paid": debt.total_paid(),
        "balance": debt.balance(),
        "pay_form": pay_form,
    })


@login_required
@role_required(["admin"])
def my_debt_add_payment(request, pk):
    debt = get_object_or_404(MyDebt, pk=pk)

    if request.method != "POST":
        return redirect("sales:my_debt_detail", pk=debt.id)

    form = MyDebtPaymentForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Fix payment errors.")
        return redirect("sales:my_debt_detail", pk=debt.id)

    amount = Decimal(form.cleaned_data["amount"] or 0)
    if amount <= 0:
        messages.error(request, "Payment must be greater than 0.")
        return redirect("sales:my_debt_detail", pk=debt.id)

    current_balance = debt.balance()
    if amount > current_balance:
        amount = current_balance
        messages.warning(request, f"Payment reduced to balance: {current_balance}")

    try:
        with transaction.atomic():
            p = form.save(commit=False)
            p.debt = debt
            p.amount = amount
            p.created_by = request.user
            p.save()

            if debt.balance() <= 0:
                debt.is_active = False
                debt.save(update_fields=["is_active", "updated_at"])

    except Exception as e:
        messages.error(request, f"Failed: {e}")
        return redirect("sales:my_debt_detail", pk=debt.id)

    messages.success(request, "Payment recorded.")
    return redirect("sales:my_debt_detail", pk=debt.id)